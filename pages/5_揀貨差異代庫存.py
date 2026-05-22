from __future__ import annotations

import io
import os
from typing import List, Dict, Tuple

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font

from common_ui import (
    inject_logistics_theme,
    set_page,
    card_open,
    card_close,
)

# =========================
# 嘗試啟用 Rich Text（若版本不支援，走 fallback）
# =========================
RICH_OK = True
try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock  # type: ignore
    from openpyxl.cell.text import InlineFont  # type: ignore
except Exception:
    RICH_OK = False


# =========================
# 通用讀檔：xlsx/xls/html假xls/csv/tsv
# =========================
def sniff_file_type_bytes(b: bytes) -> str:
    head = b[:8]
    if head[:2] == b"PK":
        return "xlsx"
    if head[:4] == b"\xD0\xCF\x11\xE0":
        return "xls"
    return "text"


def read_table_any_bytes(file_bytes: bytes, filename: str) -> pd.DataFrame:
    ftype = sniff_file_type_bytes(file_bytes)

    if ftype == "xlsx":
        return pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")

    if ftype == "xls":
        # 需要 xlrd；你原本程式就是這樣
        return pd.read_excel(io.BytesIO(file_bytes), engine="xlrd")

    # HTML 表格（ERP 假 xls 常見）
    try:
        tables = pd.read_html(io.BytesIO(file_bytes), encoding="utf-8", flavor="lxml")
        if tables and len(tables[0].columns) > 1:
            return tables[0]
    except Exception:
        pass

    # CSV/TSV 猜分隔符與編碼
    encodings = ["utf-8-sig", "cp950", "big5", "utf-8"]
    seps = ["\t", ",", ";", "|"]
    last_err = None
    for enc in encodings:
        for sep in seps:
            try:
                df = pd.read_csv(io.BytesIO(file_bytes), encoding=enc, sep=sep, engine="python")
                if df.shape[1] >= 2:
                    return df
            except Exception as e:
                last_err = e

    raise ValueError(f"❌ 無法辨識檔案為 Excel 或文字表格：{filename}\n最後錯誤：{last_err}")


# =========================
# 儲位明細：自動抓欄位 + 建立 儲位 -> 棚別 對照
# =========================
def detect_col(df: pd.DataFrame, candidates: List[str]) -> str | None:
    cols = [str(c).strip() for c in df.columns]
    for cand in candidates:
        if cand in cols:
            return cand
    # 放寬：包含關鍵字
    for cand in candidates:
        for c in cols:
            if cand in c:
                return c
    return None


def build_loc_to_shelf(slot_df: pd.DataFrame) -> Dict[str, str]:
    slot_df = slot_df.copy()
    slot_df.columns = slot_df.columns.astype(str).str.strip()

    loc_col = detect_col(slot_df, ["儲位", "儲位編號", "Location", "LOC", "Loc"])
    shelf_col = detect_col(slot_df, ["棚別", "棚架", "Shelf", "SHELF"])

    if not loc_col or not shelf_col:
        raise ValueError(
            "❌ 儲位明細抓不到必要欄位。\n"
            f"找到儲位欄：{loc_col}\n"
            f"找到棚別欄：{shelf_col}\n"
            f"目前欄位：{slot_df.columns.tolist()}"
        )

    tmp = slot_df[[loc_col, shelf_col]].copy()
    tmp[loc_col] = tmp[loc_col].astype(str).str.strip()
    tmp[shelf_col] = tmp[shelf_col].astype(str).str.strip()
    tmp = tmp[(tmp[loc_col] != "") & (tmp[loc_col].str.lower() != "nan")]

    return dict(zip(tmp[loc_col], tmp[shelf_col]))


# =========================
# 少揀檔：各自計算 -> final_df (尚未加棚別欄)
# =========================
def process_one_short(
    df_original: pd.DataFrame,
    barcode_df: pd.DataFrame,
    stock_df: pd.DataFrame,
) -> Tuple[pd.DataFrame, int]:
    """
    保留你原本邏輯：
    - RF揀貨量 >= 0
    - pivot 商品彙總：差異量 = RF - 應揀；取差異量 < 0
    - 多效期展開（每商品每效期一列）
    - 合併國際條碼
    - 庫存過濾：商品號、Canuseqty>0、排除特定儲位
    - 展開儲位/效期 pairs（儲位1/效期1…）
    """
    df_original = df_original.copy()
    df_original.columns = df_original.columns.astype(str).str.strip()

    # 你原本就假設這些欄位存在
    required = ["商品", "應揀量", "RF揀貨量", "效期"]
    miss = [c for c in required if c not in df_original.columns]
    if miss:
        raise ValueError(f"❌ 少揀明細缺少必要欄位：{miss}\n目前欄位：{df_original.columns.tolist()}")

    df_original = df_original[df_original["RF揀貨量"] >= 0].copy()

    pivot = pd.pivot_table(
        df_original,
        index="商品",
        values=["應揀量", "RF揀貨量"],
        aggfunc="sum",
    )
    pivot["差異量"] = pivot["RF揀貨量"] - pivot["應揀量"]
    underpicked = pivot[pivot["差異量"] < 0].copy()

    # 多效期展開（每種效期各一列）
    result_rows = []
    for product in underpicked.index:
        exps = df_original.loc[df_original["商品"] == product, "效期"].dropna().unique().tolist()
        if len(exps) == 0:
            result_rows.append(
                [
                    product,
                    underpicked.at[product, "應揀量"],
                    underpicked.at[product, "RF揀貨量"],
                    underpicked.at[product, "差異量"],
                    None,
                ]
            )
        elif len(exps) == 1:
            result_rows.append(
                [
                    product,
                    underpicked.at[product, "應揀量"],
                    underpicked.at[product, "RF揀貨量"],
                    underpicked.at[product, "差異量"],
                    exps[0],
                ]
            )
        else:
            for exp in sorted(exps):
                result_rows.append(
                    [
                        product,
                        underpicked.at[product, "應揀量"],
                        underpicked.at[product, "RF揀貨量"],
                        underpicked.at[product, "差異量"],
                        exp,
                    ]
                )

    result_df = pd.DataFrame(result_rows, columns=["商品", "應揀量", "RF揀貨量", "差異量", "效期"])
    result_df["效期"] = pd.to_datetime(result_df["效期"], errors="coerce").dt.strftime("%Y/%m/%d")

    # 合併國際條碼
    bc = barcode_df.copy()
    bc.columns = bc.columns.astype(str).str.strip()
    if "商品號" in bc.columns and "商品" not in bc.columns:
        bc = bc.rename(columns={"商品號": "商品"})
    if "國際條碼" in bc.columns and "商品" in bc.columns:
        result_df = result_df.merge(bc[["商品", "國際條碼"]], on="商品", how="left")

    result_df = result_df.drop_duplicates(subset=["商品", "效期"]).reset_index(drop=True)

    # 庫存過濾
    st_df = stock_df.copy()
    st_df.columns = st_df.columns.astype(str).str.strip()

    # 你原本使用這些欄位名（若不同可在你的共用檔內先修欄名）
    required_stock = ["商品號", "儲位", "Canuseqty", "商品效期"]
    miss2 = [c for c in required_stock if c not in st_df.columns]
    if miss2:
        raise ValueError(f"❌ 庫存明細缺少必要欄位：{miss2}\n目前欄位：{st_df.columns.tolist()}")

    st_df["商品效期"] = pd.to_datetime(st_df.get("商品效期"), errors="coerce").dt.strftime("%Y/%m/%d")

    excluded_locations = ["QC", "PD99", "QC99", "JCPL", "GRP", "CGS"]
    stock_filtered = st_df[
        (st_df["商品號"].isin(result_df["商品"]))
        & (st_df["Canuseqty"] > 0)
        & (~st_df["儲位"].isin(excluded_locations))
    ].copy()

    # 依「商品 + 主效期」取庫存的儲位與效期，展開成 pairs 欄位
    expanded_rows = []
    max_pairs = 0

    for _, row in result_df.iterrows():
        product = row["商品"]
        main_exp = row["效期"]

        sub = stock_filtered[stock_filtered["商品號"] == product].copy()
        # 你原本沒有強制庫存效期==主效期，只是在後面做「效期一致綠底」
        # 這裡保留：全部可用庫存都列
        pairs = list(zip(sub["儲位"].astype(str).tolist(), sub["商品效期"].astype(str).tolist()))
        max_pairs = max(max_pairs, len(pairs))

        expanded_rows.append(list(row.values) + [v for pair in pairs for v in pair])

    pair_cols = []
    for i in range(max_pairs):
        pair_cols += [f"儲位{i+1}", f"效期{i+1}"]

    final_df = pd.DataFrame(expanded_rows, columns=result_df.columns.tolist() + pair_cols)
    return final_df, max_pairs


# =========================
# 組合欄位：統一最大對數、插入棚別欄、固定欄位順序
# =========================
def normalize_and_add_shelf(df: pd.DataFrame, max_pairs: int, loc_to_shelf: Dict[str, str]) -> pd.DataFrame:
    df = df.copy()
    for i in range(1, max_pairs + 1):
        if f"儲位{i}" not in df.columns:
            df[f"儲位{i}"] = None
        if f"效期{i}" not in df.columns:
            df[f"效期{i}"] = None

    # 從後往前插入棚別欄（跟你原版一致）
    for i in range(max_pairs, 0, -1):
        loc_col = f"儲位{i}"
        shelf_col = f"棚別{i}"
        if shelf_col not in df.columns and loc_col in df.columns:
            idx = df.columns.get_loc(loc_col)
            shelves = df[loc_col].astype(str).str.strip().map(loc_to_shelf).fillna("無法對應")
            shelves = shelves.where(df[loc_col].notna(), None)
            df.insert(idx + 1, shelf_col, shelves)

    base_cols = ["商品", "應揀量", "RF揀貨量", "差異量", "效期", "國際條碼"]
    ordered = [c for c in base_cols if c in df.columns]
    for i in range(1, max_pairs + 1):
        ordered += [f"儲位{i}", f"棚別{i}", f"效期{i}"]

    rest = [c for c in df.columns if c not in ordered]
    return df[ordered + rest]


# =========================
# Excel 輔助
# =========================
def write_headers(ws, columns: List[str]):
    for c, name in enumerate(columns, start=1):
        ws.cell(row=1, column=c, value=name)


def write_df_rows(ws, df: pd.DataFrame, start_row: int) -> int:
    end_row = start_row - 1
    for r_i, row in enumerate(df.itertuples(index=False), start=start_row):
        for c_i, val in enumerate(row, start=1):
            ws.cell(row=r_i, column=c_i, value=val)
        end_row = r_i
    return end_row


def find_col_index(ws, header_name: str) -> int | None:
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == header_name:
            return c
    return None


def normalize_barcode_value(v) -> str:
    """把可能是數字/浮點/科學記號的條碼安全轉成字串，並盡量補 13 碼"""
    if v is None:
        return ""
    s = str(v).strip()
    if s == "" or s.lower() == "nan":
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    if s.isdigit() and len(s) < 13:
        s = s.zfill(13)
    return s


def apply_barcode_last5_big_all(ws, barcode_col: int, start_row: int = 2) -> bool:
    """
    全欄套用：國際條碼同一格後五碼放大（RichText）。
    True=成功；False=不支援或失敗。
    """
    if not RICH_OK:
        return False

    normal_f = InlineFont(sz=11)  # type: ignore
    big_f = InlineFont(sz=16, b=True)  # type: ignore

    try:
        for r in range(start_row, ws.max_row + 1):
            ws.cell(row=r, column=barcode_col).number_format = "@"

        for r in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=r, column=barcode_col)
            s = normalize_barcode_value(cell.value)
            if not s:
                continue

            last5 = s[-5:] if len(s) >= 5 else s
            prefix = s[:-5] if len(s) >= 5 else ""

            rich = CellRichText()  # type: ignore
            if prefix:
                rich.append(TextBlock(normal_f, prefix))  # type: ignore
            rich.append(TextBlock(big_f, last5))  # type: ignore
            cell.value = rich

        return True
    except Exception:
        return False


def add_barcode_last5_column_fallback(ws, barcode_col: int):
    """
    若 RichText 不支援：新增『國際條碼_後五碼』欄並整格放大。
    """
    insert_pos = barcode_col + 1
    ws.insert_cols(insert_pos)
    ws.cell(row=1, column=insert_pos, value="國際條碼_後五碼")

    big_font = Font(size=16, bold=True)
    for r in range(2, ws.max_row + 1):
        s = normalize_barcode_value(ws.cell(row=r, column=barcode_col).value)
        last5 = s[-5:] if len(s) >= 5 else s
        c = ws.cell(row=r, column=insert_pos, value=last5)
        c.font = big_font


def build_export_xlsx_bytes(dfs2: List[pd.DataFrame], output_sheet_name: str = "結果") -> Tuple[bytes, str]:
    """
    - 單一工作表
    - 多檔接續貼上（不留空白行）
    - 綠底：效期2/效期3… == 主效期
    - 黃底：第二份少揀檔區塊只黃國際條碼欄
    - 國際條碼後五碼放大（RichText 或 fallback 欄）
    """
    if not dfs2:
        raise ValueError("❌ 沒有可輸出的資料")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = output_sheet_name

    columns = list(dfs2[0].columns)
    write_headers(ws, columns)

    blocks = []
    cur_row = 2
    for idx, df in enumerate(dfs2, start=1):
        start = cur_row
        end = write_df_rows(ws, df, start_row=start)
        blocks.append({"idx": idx, "start": start, "end": end})
        cur_row = end + 1

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    main_exp_col = find_col_index(ws, "效期")
    barcode_col = find_col_index(ws, "國際條碼")
    if main_exp_col is None:
        raise ValueError("❌ 找不到主欄位『效期』，請確認輸出表頭")
    if barcode_col is None:
        raise ValueError("❌ 找不到欄位『國際條碼』，請確認輸出表頭")

    exp_cols = []
    for c in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=c).value
        if isinstance(header, str) and header.startswith("效期") and header != "效期":
            exp_cols.append(c)

    # 標色：效期一致 -> 綠底（欄位：效期2/效期3...）
    for r in range(2, ws.max_row + 1):
        ref = ws.cell(row=r, column=main_exp_col).value
        if ref in (None, "", "nan"):
            continue
        for c in exp_cols:
            if ws.cell(row=r, column=c).value == ref:
                ws.cell(row=r, column=c).fill = green_fill

    # 全欄後五碼放大
    used_rich = apply_barcode_last5_big_all(ws, barcode_col, start_row=2)
    if not used_rich:
        add_barcode_last5_column_fallback(ws, barcode_col)

    # 第二份：只黃國際條碼欄
    block2 = next((b for b in blocks if b["idx"] == 2), None)
    if block2:
        # 若 fallback 插入了新欄，barcode_col 位置仍是原本的國際條碼欄（沒變）
        for r in range(block2["start"], block2["end"] + 1):
            ws.cell(row=r, column=barcode_col).fill = yellow_fill

    out = io.BytesIO()
    wb.save(out)
    wb.close()

    note = "RichText" if used_rich else "Fallback(新增後五碼欄)"
    return out.getvalue(), note


# =========================
# Streamlit Page
# =========================
def main():
    inject_logistics_theme()
    set_page("揀貨差異分析（庫存定位強化）", icon="🔎", subtitle="少揀差異｜庫存儲位展開｜棚別對應｜國際條碼後五碼放大")

    with st.sidebar:
        st.header("⚙️ 設定")
        st.caption("依原邏輯：少揀可多檔上傳，會依序接續輸出到同一張 Sheet")
        show_preview_rows = st.number_input("預覽筆數", min_value=50, max_value=5000, value=500, step=50)

    card_open("📤 檔案上傳")
    short_files = st.file_uploader(
        "少揀明細（可多選）",
        type=["xlsx", "xls", "xlsm", "csv", "txt"],
        accept_multiple_files=True,
    )
    common_file = st.file_uploader(
        "商品對照表 / 庫存明細（同一個檔）",
        type=["xlsx", "xls", "xlsm", "csv", "txt"],
        accept_multiple_files=False,
    )
    slot_file = st.file_uploader(
        "儲位明細（含棚別）",
        type=["xlsx", "xls", "xlsm", "csv", "txt"],
        accept_multiple_files=False,
    )

    run = st.button("🚀 產出分析", type="primary", disabled=(not short_files or not common_file or not slot_file))
    card_close()

    if not run:
        st.info("請依序上傳：少揀明細（可多檔）＋ 商品對照/庫存明細（同檔）＋ 儲位明細（含棚別）")
        return

    try:
        with st.spinner("資料處理中..."):
            # 共同檔：同一檔讀兩次（保留你原本邏輯）
            common_xlsx = read_table_any_bytes(common_file.getvalue(), common_file.name)
            barcode_df = common_xlsx.copy()
            stock_df = common_xlsx.copy()

            # 國際條碼補 13 碼（空值保留空）
            barcode_df.columns = barcode_df.columns.astype(str).str.strip()
            if "國際條碼" in barcode_df.columns:
                barcode_df["國際條碼"] = barcode_df["國際條碼"].apply(
                    lambda x: str(x).zfill(13) if pd.notna(x) and str(x).strip() != "" else ""
                )

            # 儲位明細 -> 儲位:棚別 map
            slot_df = read_table_any_bytes(slot_file.getvalue(), slot_file.name)
            loc_to_shelf = build_loc_to_shelf(slot_df)

            # 每份少揀檔各自算出 df（不合併計算）
            dfs = []
            max_pairs_global = 0

            for f in short_files:
                df_original = read_table_any_bytes(f.getvalue(), f.name)
                final_df, max_pairs = process_one_short(df_original, barcode_df, stock_df)
                # 預覽/辨識用：加批次與檔名（不影響你原本運算）
                final_df.insert(0, "批次", os.path.basename(f.name))
                dfs.append(final_df)
                max_pairs_global = max(max_pairs_global, max_pairs)

            # 統一欄位 + 插入棚別 + 固定順序
            dfs2 = []
            for df in dfs:
                # normalize_and_add_shelf 期望 base_cols 開頭是商品...；這裡批次在最前面，所以先暫存再插回
                batch = df["批次"].copy()
                df2 = df.drop(columns=["批次"]).copy()
                df2 = normalize_and_add_shelf(df2, max_pairs_global, loc_to_shelf)
                df2.insert(0, "批次", batch.values)
                dfs2.append(df2)

            # 合併預覽（畫面用）
            preview_df = pd.concat(dfs2, ignore_index=True)

            # 匯出 bytes
            xlsx_bytes, mode_note = build_export_xlsx_bytes(dfs2, output_sheet_name="結果")

        card_open("🧾 結果預覽")
        st.caption(f"匯出處理：國際條碼後五碼放大模式 = {mode_note}")
        st.dataframe(preview_df.head(int(show_preview_rows)), use_container_width=True, hide_index=True)
        card_close()

        st.download_button(
            "⬇️ 匯出報表（Excel）",
            data=xlsx_bytes,
            file_name="揀貨差異_多檔接續輸出_含棚別_後五碼放大.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error("❌ 執行失敗")
        st.code(str(e))


if __name__ == "__main__":
    main()
