# -*- coding: utf-8 -*-
import os
import re
import math
import hashlib
import base64
import pandas as pd
from io import BytesIO

import streamlit as st
import streamlit.components.v1 as components

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

# 你的平台若有 common_ui 就用；沒有也可直接註解掉
try:
    from common_ui import inject_logistics_theme, set_page, card_open, card_close
except Exception:
    inject_logistics_theme = lambda: None
    set_page = lambda *a, **k: None
    card_open = lambda *a, **k: None
    card_close = lambda *a, **k: None


# =========================
# 基本設定
# =========================
TARGET_PER_MANHOUR = 790
AM_HOURS = list(range(8, 13))     # 08-12
PM_HOURS = list(range(13, 19))    # 13-18

BASE_FONT_NAME = "微軟正黑體"
BASE_FONT_SIZE = 12
ROW_HEIGHT = 18

NUM_FMT_2_HIDE0 = "#,##0.00;-#,##0.00;;@"
NUM_FMT_4_HIDE0 = "#,##0.0000;-#,##0.0000;;@"
NUM_FMT_MAN_INT_HIDE0 = "#,##0;-#,##0;;@"
NUM_FMT_MAN_FLOAT_HIDE0 = "#,##0.##;-#,##0.##;;@"
NUM_FMT_INT_HIDE0 = "#,##0;-#,##0;;@"
NUM_FMT_INT = "#,##0"
NUM_FMT_MONEY_HIDE0 = "#,##0;-#,##0;;@"

AM_DEFAULT_MONEY = 100
PM_DEFAULT_MONEY = 50

QCA_LIST = ["GT-B", "GT-C", "GT-D", "GT-E"]
QCB_LIST = ["GT-A", "GT-J", "GT-K"]


# =========================
# 讀檔（支援假xls/HTML/CSV）
# =========================
OLE_HEADER = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"
ZIP_HEADER = b"PK\x03\x04"


def _try_read_html(raw: bytes) -> pd.DataFrame:
    tables = pd.read_html(BytesIO(raw))
    if tables:
        return tables[0]
    raise ValueError("HTML 讀取不到表格")


def _try_read_text_like(raw: bytes) -> pd.DataFrame:
    for enc in ("utf-8-sig", "utf-8", "big5", "cp950", "latin1"):
        for sep in ("\t", ",", ";", "|"):
            try:
                df = pd.read_csv(BytesIO(raw), encoding=enc, sep=sep)
                if df.shape[1] >= 2:
                    return df
            except Exception:
                continue
    raise ValueError("不是可解析的文字分隔檔")


def robust_read_bytes(raw: bytes, filename: str) -> pd.DataFrame:
    ext = os.path.splitext(filename)[1].lower()
    head = raw[:8]
    is_ole = head.startswith(OLE_HEADER)
    is_zip = head.startswith(ZIP_HEADER)

    if is_zip or ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return pd.read_excel(BytesIO(raw), engine="openpyxl")

    if is_ole:
        try:
            return pd.read_excel(BytesIO(raw), engine="xlrd")
        except Exception:
            try:
                return pd.read_excel(BytesIO(raw), engine="openpyxl")
            except Exception:
                try:
                    return _try_read_html(raw)
                except Exception:
                    return _try_read_text_like(raw)

    if ext == ".csv":
        for enc in ("utf-8-sig", "utf-8", "big5", "cp950", "latin1"):
            try:
                return pd.read_csv(BytesIO(raw), encoding=enc)
            except Exception:
                continue
        return pd.read_csv(BytesIO(raw), encoding="utf-8", errors="replace")

    try:
        return _try_read_html(raw)
    except Exception:
        return _try_read_text_like(raw)


# =========================
# Excel 字體/列高工具
# =========================
def _clone_font(cell_font: Font, *, name=None, size=None, bold=None, color=None):
    if cell_font is None:
        cell_font = Font()
    return Font(
        name=name if name is not None else cell_font.name,
        size=size if size is not None else cell_font.size,
        bold=bold if bold is not None else cell_font.bold,
        italic=cell_font.italic,
        vertAlign=cell_font.vertAlign,
        underline=cell_font.underline,
        strike=cell_font.strike,
        color=color if color is not None else cell_font.color,
        outline=cell_font.outline,
        shadow=cell_font.shadow,
        condense=cell_font.condense,
        extend=cell_font.extend,
        charset=cell_font.charset,
        family=cell_font.family,
        scheme=cell_font.scheme,
    )


def _set_base_font(cell, *, force_color=None, force_bold=None):
    cell.font = _clone_font(
        cell.font,
        name=BASE_FONT_NAME,
        size=BASE_FONT_SIZE,
        bold=force_bold if force_bold is not None else cell.font.bold,
        color=force_color if force_color is not None else cell.font.color,
    )


def _set_row_height(ws, r):
    ws.row_dimensions[r].height = ROW_HEIGHT


# =========================
# 欄位對照
# =========================
def normalize_columns(df: pd.DataFrame):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    colmap = {str(c).strip().upper(): c for c in df.columns}

    def pick(*cands):
        for k in cands:
            if k in colmap:
                return colmap[k]
        return None

    c_pickdate = pick("PICKDATE", "PICK_DATE", "PICK DATETIME", "PICKTIME", "PICK_TIME")
    c_packqty  = pick("PACKQTY", "PACK_QTY", "PCS", "QTY")
    c_cweight  = pick("CWEIGHT", "C_WEIGHT", "C WEIGHT", "WEIGHT")
    c_lineid   = pick("LINEID", "LINE_ID", "LINE", "LINE ID")
    c_stotype  = pick("STO_TYPE", "STOTYPE", "SO_TYPE", "TYPE")

    missing = [name for name, col in [
        ("PICKDATE", c_pickdate),
        ("PACKQTY",  c_packqty),
        ("Cweight",  c_cweight),
        ("LINEID",   c_lineid),
        ("STO_TYPE", c_stotype),
    ] if col is None]

    if missing:
        raise KeyError(f"缺少必要欄位：{missing}\n目前欄位：{list(df.columns)}")

    return df, c_pickdate, c_packqty, c_cweight, c_lineid, c_stotype


# =========================
# 建立每小時彙整
# =========================
def build_hourly_metrics(df, c_pickdate, c_packqty, c_cweight, c_lineid, c_stotype):
    df = df.copy()
    df[c_stotype] = df[c_stotype].astype(str).str.strip()
    df[c_lineid] = df[c_lineid].astype(str).str.strip()

    df[c_pickdate] = pd.to_datetime(df[c_pickdate], errors="coerce")
    df[c_packqty] = pd.to_numeric(df[c_packqty], errors="coerce").fillna(0)
    df[c_cweight] = pd.to_numeric(df[c_cweight], errors="coerce").fillna(0)

    df["加權PCS"] = df[c_packqty] * df[c_cweight]
    df["PICK_HOUR"] = df[c_pickdate].dt.floor("h")
    df["PICK_DATE"] = df["PICK_HOUR"].dt.date
    df["HOUR"] = df["PICK_HOUR"].dt.hour

    line_base = (df.groupby(["PICK_DATE", c_lineid, "HOUR"])[[c_packqty, "加權PCS"]]
                   .sum()
                   .reset_index()
                   .rename(columns={c_packqty: "PCS"}))

    split = (df.groupby(["PICK_DATE", c_lineid, "HOUR", c_stotype])[[c_packqty, "加權PCS"]]
               .sum()
               .reset_index()
               .rename(columns={c_packqty: "PCS"}))

    return df, line_base, split


# =========================
# 名單：讓使用者貼上（TSV）
# =========================
DEFAULT_NAME_TSV = """補貨\t邱清瑞\t楊文點\t阮功水
GT-B\t范明俊\t黃口康\t吳黃金珠\t陳先權
GT-C\t潘文一\t郭雙燕\t黎金妮\t廖永成
GT-D\t蔡麗珠\t阮瑞美黃綠\t阮黃英\t岳子恆
GT-E\t阮玉名\t潘氏青江\t王文楷\t李杰儒
GT-J\t阮伊黃\t柴家欣\t黎氏瑋\t阮孟勇
GT-K\t河文強\t楊心如\t阮氏美麗\t楊浩傑
GT-A\t李茂銓\t陳國慶\t阮武玉玄\t潘氏慶平
QCA\t葉欲弘\t范文春\t嚴彩康
QCB\t阮文忠\t何玉進
"""


def parse_name_tsv(tsv_text: str) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {}
    if not tsv_text:
        return out
    for raw_line in tsv_text.splitlines():
        line = raw_line.strip("\n\r")
        if not line.strip():
            continue
        parts = re.split(r"\t|,|\s{2,}", line.strip())
        parts = [p.strip() for p in parts if p.strip()]
        if len(parts) < 2:
            continue
        key = parts[0]
        names = parts[1:]
        out[key] = names
    return out


def names_for_line(line_id: str, name_map: dict[str, list[str]]) -> str:
    line_id = (line_id or "").strip()
    names = name_map.get(line_id, [])
    names = [n for n in names if str(n).strip()]
    return "、".join(names) if names else ""


# =========================
# manpower helpers
# =========================
def _as_float(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def _manpower_cell_value_and_format(raw):
    if raw is None or raw == "":
        return "", None
    try:
        v = float(raw)
    except Exception:
        return "", None
    if abs(v) < 1e-12:
        return "", None
    if abs(v - round(v)) < 1e-12:
        return int(round(v)), NUM_FMT_MAN_INT_HIDE0
    return float(v), NUM_FMT_MAN_FLOAT_HIDE0


# =========================
# 人力表格：建立/對齊（排除變更不會清空）
# =========================
def build_default_manpower_table(lineids, hours):
    cols = ["Line ID"] + [str(int(h)) for h in hours]
    data = {"Line ID": list(lineids)}
    for h in hours:
        data[str(int(h))] = [""] * len(lineids)
    return pd.DataFrame(data, columns=cols)


def reconcile_manpower_table(existing: pd.DataFrame, lineids, hours) -> pd.DataFrame:
    new_df = build_default_manpower_table(lineids, hours)

    if existing is None or existing.empty:
        return new_df

    keep_cols = ["Line ID"] + [str(int(h)) for h in hours]
    existing_cols = [c for c in keep_cols if c in existing.columns]
    tmp = existing[existing_cols].copy()

    tmp["Line ID"] = tmp["Line ID"].astype(str).str.strip()
    tmp = tmp.set_index("Line ID")

    new_df["Line ID"] = new_df["Line ID"].astype(str).str.strip()
    new_df = new_df.set_index("Line ID")

    for lid in new_df.index:
        if lid in tmp.index:
            for c in new_df.columns:
                if c in tmp.columns:
                    new_df.loc[lid, c] = tmp.loc[lid, c]

    return new_df.reset_index()


def manpower_table_to_map(d, table_df: pd.DataFrame):
    mp = {}
    if table_df is None or table_df.empty:
        return mp
    if "Line ID" not in table_df.columns:
        return mp

    hours_cols = [c for c in table_df.columns if c != "Line ID"]

    for _, row in table_df.iterrows():
        lid = str(row["Line ID"]).strip()
        if not lid:
            continue
        for hc in hours_cols:
            try:
                hour = int(str(hc).strip())
            except Exception:
                continue
            v = row.get(hc, "")
            if v is None or (isinstance(v, float) and pd.isna(v)):
                v = ""
            if isinstance(v, str):
                v = v.strip()
            mp[(d, lid, hour)] = v

    return mp


def build_line_base_map_for_date(d, lineids, line_base, c_lineid):
    sub_base = line_base[line_base["PICK_DATE"] == d]
    line_base_map = {}
    for lid in lineids:
        tmp = sub_base[sub_base[c_lineid].astype(str).str.strip() == str(lid)]
        line_base_map[(str(lid), "PCS")] = {int(r["HOUR"]): r["PCS"] for _, r in tmp.iterrows()}
        line_base_map[(str(lid), "加權PCS")] = {int(r["HOUR"]): r["加權PCS"] for _, r in tmp.iterrows()}
    return line_base_map


# =========================
# 達標狀況（即時顯示；人力 2 位）
# =========================
def build_achv_status_table(date_value, lineids, hours, line_base_map, manpower_map, name_map) -> pd.DataFrame:
    rows = []
    for lid in lineids:
        lid = str(lid)
        pcs_w_map = line_base_map.get((lid, "加權PCS"), {})

        am_man = sum(_as_float(manpower_map.get((date_value, lid, int(h)), "")) for h in AM_HOURS if h in hours)
        pm_man = sum(_as_float(manpower_map.get((date_value, lid, int(h)), "")) for h in PM_HOURS if h in hours)

        am_pcs = sum(float(pcs_w_map.get(int(h), 0) or 0) for h in AM_HOURS if h in hours)
        pm_pcs = sum(float(pcs_w_map.get(int(h), 0) or 0) for h in PM_HOURS if h in hours)

        am_target = math.trunc(TARGET_PER_MANHOUR * am_man)
        pm_target = math.trunc(TARGET_PER_MANHOUR * pm_man)

        am_ok = (am_man > 0 and math.trunc(am_pcs) >= am_target)
        pm_ok = (pm_man > 0 and math.trunc(pm_pcs) >= pm_target)

        am_man_show = None if abs(am_man) < 1e-12 else round(am_man, 2)
        pm_man_show = None if abs(pm_man) < 1e-12 else round(pm_man, 2)

        rows.append({
            "Line ID": lid,
            "姓名": names_for_line(lid, name_map),
            "上午人力": am_man_show,
            "上午PCS(加權)": None if abs(am_pcs) < 1e-12 else math.trunc(am_pcs),
            "上午目標(加權)": None if abs(am_man) < 1e-12 else am_target,
            "上午達標": ("✅" if am_ok else "❌"),
            "下午人力": pm_man_show,
            "下午PCS(加權)": None if abs(pm_pcs) < 1e-12 else math.trunc(pm_pcs),
            "下午目標(加權)": None if abs(pm_man) < 1e-12 else pm_target,
            "下午達標": ("✅" if pm_ok else "❌"),
        })

    return pd.DataFrame(rows)


def style_achv(df: pd.DataFrame):
    def _bg(v):
        if v == "✅":
            return "background-color: #C6EFCE; color: #006100; font-weight: 700;"
        if v == "❌":
            return "background-color: #FFC7CE; color: #9C0006; font-weight: 700;"
        return ""

    sty = df.style
    if "上午達標" in df.columns:
        sty = sty.applymap(_bg, subset=["上午達標"])
    if "下午達標" in df.columns:
        sty = sty.applymap(_bg, subset=["下午達標"])

    fmt = {}
    if "上午人力" in df.columns:
        fmt["上午人力"] = "{:.2f}"
    if "下午人力" in df.columns:
        fmt["下午人力"] = "{:.2f}"
    if fmt:
        sty = sty.format(fmt, na_rep="")

    return sty


# =========================
# Excel（日期表）
# =========================
def write_hourly_sheet(
    wb, sheet_name, date_value, hours, lineids,
    line_base_map, split_map, manpower_map
):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(sheet_name)

    black_fill = PatternFill("solid", fgColor="111111")
    head_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    _set_row_height(ws, 1)
    ws["A1"] = str(date_value)
    ws["A1"].alignment = center
    ws["A1"].fill = head_fill
    ws["A1"].border = border
    _set_base_font(ws["A1"], force_bold=True)

    for j, h in enumerate(hours, start=2):
        c = ws.cell(row=1, column=j, value=int(h))
        c.alignment = center
        c.fill = head_fill
        c.border = border
        _set_base_font(c, force_bold=True)

    _set_row_height(ws, 2)
    ws["A2"] = "撿貨（已撿數量）"
    ws["A2"].fill = black_fill
    ws["A2"].alignment = left
    ws["A2"].border = border
    _set_base_font(ws["A2"], force_bold=True, force_color="FFFFFF")

    for j in range(2, 2 + len(hours)):
        cell = ws.cell(row=2, column=j, value=None)
        cell.alignment = right
        cell.border = border
        cell.number_format = NUM_FMT_2_HIDE0
        _set_base_font(cell)

    r = 3
    pcs_weight_rows = []

    def write_row(label, values_by_hour=None, fill=None, is_manpower=False):
        nonlocal r
        _set_row_height(ws, r)

        a = ws.cell(row=r, column=1, value=label)
        a.fill = black_fill
        a.alignment = left
        a.border = border
        _set_base_font(a, force_bold=True, force_color="FFFFFF")

        for j, h in enumerate(hours, start=2):
            raw = "" if values_by_hour is None else values_by_hour.get(int(h), "")

            if is_manpower:
                val, fmt = _manpower_cell_value_and_format(raw)
                c = ws.cell(row=r, column=j, value=val)
                if fmt:
                    c.number_format = fmt
            else:
                try:
                    fv = float(raw)
                    if abs(fv) < 1e-12:
                        c = ws.cell(row=r, column=j, value="")
                    else:
                        c = ws.cell(row=r, column=j, value=fv)
                        c.number_format = NUM_FMT_2_HIDE0
                except Exception:
                    c = ws.cell(row=r, column=j, value="")

            c.alignment = right
            c.border = border
            if fill is not None:
                c.fill = fill
                _set_base_font(c, force_bold=True)
            else:
                _set_base_font(c)

        row_idx = r
        r += 1
        return row_idx

    def write_formula_row(label, numerator_row, denom_row, number_format):
        nonlocal r
        _set_row_height(ws, r)

        a = ws.cell(row=r, column=1, value=label)
        a.fill = black_fill
        a.alignment = left
        a.border = border
        _set_base_font(a, force_bold=True, force_color="FFFFFF")

        for j in range(2, 2 + len(hours)):
            col = get_column_letter(j)
            num = f"{col}{numerator_row}"
            den = f"{col}{denom_row}"
            formula = f'=IF(OR({den}="",{den}=0),"",IF({num}/{den}=0,"",{num}/{den}))'
            c = ws.cell(row=r, column=j, value=formula)
            c.alignment = right
            c.border = border
            c.number_format = number_format
            _set_base_font(c)

        r += 1

    for lid in lineids:
        lid = str(lid)

        _set_row_height(ws, r)
        a = ws.cell(row=r, column=1, value=f"{lid} Line")
        a.fill = black_fill
        a.alignment = left
        a.border = border
        _set_base_font(a, force_bold=True, force_color="FFFFFF")
        for j in range(2, 2 + len(hours)):
            c = ws.cell(row=r, column=j, value=None)
            c.border = border
            c.alignment = right
            _set_base_font(c)
        r += 1

        pcs_weight_map = line_base_map.get((lid, "加權PCS"), {})
        pcs_map = line_base_map.get((lid, "PCS"), {})

        gso_w = split_map.get((lid, "GSO", "加權PCS"), {})
        gxso_w = split_map.get((lid, "GXSO", "加權PCS"), {})
        gso = split_map.get((lid, "GSO", "PCS"), {})
        gxso = split_map.get((lid, "GXSO", "PCS"), {})

        row_pcs_w = write_row(f"{lid}（PCS）加權", pcs_weight_map)
        pcs_weight_rows.append(row_pcs_w)

        write_row(f"{lid}（PCS）", pcs_map)
        write_row("GSO(加權)", gso_w)
        write_row("GXSO(加權)", gxso_w)
        write_row("GSO", gso)
        write_row("GXSO", gxso)

        man_map = {int(h): manpower_map.get((date_value, lid, int(h)), "") for h in hours}
        row_man = write_row(
            f"{lid}（人數）",
            man_map,
            fill=PatternFill("solid", fgColor="FFF2CC"),
            is_manpower=True
        )

        write_formula_row("平均產力(加權) 4", numerator_row=row_pcs_w, denom_row=row_man, number_format=NUM_FMT_4_HIDE0)
        write_formula_row("平均產力(加權)", numerator_row=row_pcs_w, denom_row=row_man, number_format=NUM_FMT_2_HIDE0)

        _set_row_height(ws, r)
        for j in range(1, 2 + len(hours)):
            c = ws.cell(row=r, column=j, value=None)
            c.border = border
            c.alignment = right if j >= 2 else left
            _set_base_font(c)
        r += 1

    for j in range(2, 2 + len(hours)):
        col = get_column_letter(j)
        refs = ",".join([f"{col}{rr}" for rr in pcs_weight_rows]) if pcs_weight_rows else ""
        c = ws.cell(row=2, column=j, value=(f'=IF(SUM({refs})=0,"",SUM({refs}))' if refs else '""'))
        c.number_format = NUM_FMT_2_HIDE0
        c.alignment = Alignment(horizontal="right", vertical="center")
        _set_base_font(c)

    ws.column_dimensions["A"].width = 24
    for j in range(2, 2 + len(hours)):
        ws.column_dimensions[get_column_letter(j)].width = 12
    ws.freeze_panes = "B3"
    return ws


# =========================
# Excel：彙總 helpers
# =========================
def _is_hour(v):
    try:
        iv = int(str(v).strip())
        return 0 <= iv <= 23
    except Exception:
        return False


def find_hour_header_row(ws):
    max_r = min(ws.max_row, 10)
    max_c = min(ws.max_column, 120)
    for r in range(1, max_r + 1):
        hour_to_col = {}
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if _is_hour(v):
                hour_to_col[int(str(v).strip())] = c
        if len(hour_to_col) >= 3:
            return r, hour_to_col
    return None, {}


def parse_line_rows(ws):
    out = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not v:
            continue
        t = str(v).strip()

        m1 = re.match(r"(.+?)（PCS）加權$", t)
        if m1:
            lid = m1.group(1).strip()
            out.setdefault(lid, {})["pcs_w"] = r
            continue

        m2 = re.match(r"(.+?)（人數）$", t)
        if m2:
            lid = m2.group(1).strip()
            out.setdefault(lid, {})["man"] = r
            continue

    return {k: v for k, v in out.items() if "pcs_w" in v and "man" in v}


def sum_cells_formula(sheet, row, cols):
    if not cols:
        return "=0"
    refs = [f"'{sheet}'!{get_column_letter(c)}{row}" for c in cols]
    return f"=SUM({','.join(refs)})"


def countif_sum(range_a1: str, items: list[str]) -> str:
    return "+".join([f'COUNTIF({range_a1},"{it}")' for it in items]) if items else "0"


def compute_achievers(date_value, lineids, hours, line_base_map, manpower_map):
    am_ach, pm_ach = [], []
    for lid in lineids:
        lid = str(lid)

        am_man = sum(_as_float(manpower_map.get((date_value, lid, int(h)), "")) for h in AM_HOURS if h in hours)
        pm_man = sum(_as_float(manpower_map.get((date_value, lid, int(h)), "")) for h in PM_HOURS if h in hours)

        pcs_w_map = line_base_map.get((lid, "加權PCS"), {})
        am_pcs = sum(float(pcs_w_map.get(int(h), 0) or 0) for h in AM_HOURS if h in hours)
        pm_pcs = sum(float(pcs_w_map.get(int(h), 0) or 0) for h in PM_HOURS if h in hours)

        am_target = math.trunc(TARGET_PER_MANHOUR * am_man)
        pm_target = math.trunc(TARGET_PER_MANHOUR * pm_man)

        if am_man > 0 and math.trunc(am_pcs) >= am_target:
            am_ach.append(lid)
        if pm_man > 0 and math.trunc(pm_pcs) >= pm_target:
            pm_ach.append(lid)

    return am_ach, pm_ach


def build_summary_sheet_with_achievers(
    wb, summary_name, date_ws, line_map, am_cols, pm_cols,
    am_achievers, pm_achievers, name_map
):
    if summary_name in wb.sheetnames:
        wb.remove(wb[summary_name])
    ws = wb.create_sheet(summary_name)

    thin = Side(style="thin", color="D0D0D0")
    thick = Side(style="medium", color="111111")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    box_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    header_fill = PatternFill("solid", fgColor="F8CBAD")
    sub_header_fill = PatternFill("solid", fgColor="F2F2F2")

    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    headers = {
        1: "Line ID",
        2: "每小時目標(加權)",
        3: "上午人力",
        4: "上午應達成(加權)",
        5: "總PCS(加權)",
        6: "差異(加權)",
        7: "下午人力",
        8: "下午目標(加權)",
        9: "總PCS(加權)",
        10: "差異(加權)",
    }

    def _set_row_h(r):
        ws.row_dimensions[r].height = ROW_HEIGHT

    _set_row_h(1)
    for col, title in headers.items():
        cell = ws.cell(1, col, title)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center
        _set_base_font(cell, force_bold=True)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 2
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 28
    ws.column_dimensions["N"].width = 14

    r0 = 2
    src = date_ws.title

    def text_no_trailing_dot(expr: str):
        t = f'TEXT({expr},"#,##0.##")'
        return f'IF(RIGHT({t},1)=".",LEFT({t},LEN({t})-1),{t})'

    items = list(line_map.items())
    for i, (lid, rows) in enumerate(items):
        r = r0 + i
        _set_row_h(r)

        for col in [1,2,3,4,5,6,7,8,9,10]:
            cell = ws.cell(r, col)
            cell.border = border
            cell.alignment = center if col in (1,2,3) else right
            _set_base_font(cell)

        ws.cell(r, 1, str(lid).strip())
        ws.cell(r, 2, TARGET_PER_MANHOUR).number_format = NUM_FMT_INT

        man_row = rows["man"]
        pcsw_row = rows["pcs_w"]

        am_man_expr = sum_cells_formula(src, man_row, am_cols)[1:]
        am_pcs_expr = sum_cells_formula(src, pcsw_row, am_cols)[1:]
        pm_man_expr = sum_cells_formula(src, man_row, pm_cols)[1:]
        pm_pcs_expr = sum_cells_formula(src, pcsw_row, pm_cols)[1:]

        ws.cell(r, 3, f'=IF({am_man_expr}=0,"",{text_no_trailing_dot(am_man_expr)})').number_format = "@"
        ws.cell(r, 7, f'=IF({pm_man_expr}=0,"",{text_no_trailing_dot(pm_man_expr)})').number_format = "@"

        ws.cell(r, 4, f'=IF($C{r}="","",TRUNC($B{r}*VALUE(SUBSTITUTE($C{r},",","")),0))').number_format = NUM_FMT_INT_HIDE0
        ws.cell(r, 8, f'=IF($G{r}="","",TRUNC($B{r}*VALUE(SUBSTITUTE($G{r},",","")),0))').number_format = NUM_FMT_INT_HIDE0

        ws.cell(r, 5, f'=IF($C{r}="","",IF({am_pcs_expr}=0,"",TRUNC({am_pcs_expr},0)))').number_format = NUM_FMT_INT_HIDE0
        ws.cell(r, 9, f'=IF($G{r}="","",IF({pm_pcs_expr}=0,"",TRUNC({pm_pcs_expr},0)))').number_format = NUM_FMT_INT_HIDE0

        ws.cell(r, 6, f'=IF(OR($D{r}="",$E{r}=""),"",TRUNC($E{r}-$D{r},0))').number_format = NUM_FMT_INT_HIDE0
        ws.cell(r, 10, f'=IF(OR($H{r}="",$I{r}=""),"",TRUNC($I{r}-$H{r},0))').number_format = NUM_FMT_INT_HIDE0

    last_row = r0 + len(items) - 1
    if last_row < r0:
        last_row = r0

    if last_row >= r0 and items:
        red_fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
        green_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
        dxf_red = DifferentialStyle(font=Font(color="9C0006"), fill=red_fill)
        dxf_green = DifferentialStyle(font=Font(color="006100"), fill=green_fill)

        rule_e_red = Rule(type="expression", dxf=dxf_red, stopIfTrue=True)
        rule_e_red.formula = [f'AND($D{r0}<>"",E{r0}<>"",E{r0}<$D{r0})']
        ws.conditional_formatting.add(f"E{r0}:E{last_row}", rule_e_red)

        rule_e_green = Rule(type="expression", dxf=dxf_green, stopIfTrue=True)
        rule_e_green.formula = [f'AND($D{r0}<>"",E{r0}<>"",E{r0}>=$D{r0})']
        ws.conditional_formatting.add(f"E{r0}:E{last_row}", rule_e_green)

        rule_i_red = Rule(type="expression", dxf=dxf_red, stopIfTrue=True)
        rule_i_red.formula = [f'AND($H{r0}<>"",I{r0}<>"",I{r0}<$H{r0})']
        ws.conditional_formatting.add(f"I{r0}:I{last_row}", rule_i_red)

        rule_i_green = Rule(type="expression", dxf=dxf_green, stopIfTrue=True)
        rule_i_green.formula = [f'AND($H{r0}<>"",I{r0}<>"",I{r0}>=$H{r0})']
        ws.conditional_formatting.add(f"I{r0}:I{last_row}", rule_i_green)

    # ✅ 無達標：不顯示 QCA/QCB/補貨
    def draw_achiever_block(start_row, title, achievers, default_money, qc_mult, restock_mult):
        col_line = 12  # L
        col_name = 13  # M
        col_money = 14 # N

        _set_row_h(start_row)
        for cc in range(col_line, col_money + 1):
            c = ws.cell(start_row, cc)
            c.border = border
            _set_base_font(c)

        tcell = ws.cell(start_row, col_line, title)
        tcell.alignment = left
        _set_base_font(tcell, force_bold=True)
        ws.merge_cells(start_row=start_row, start_column=col_line, end_row=start_row, end_column=col_money)

        hr = start_row + 1
        _set_row_h(hr)
        h1 = ws.cell(hr, col_line, "達標 Line ID")
        h2 = ws.cell(hr, col_name, "姓名（自動帶入，可手改）")
        h3 = ws.cell(hr, col_money, "金額（可輸入/預設公式）")
        for c in (h1, h2, h3):
            c.fill = sub_header_fill
            c.border = border
            c.alignment = center
            _set_base_font(c, force_bold=True)

        rr = hr + 1
        first_list_row = rr

        def _write_row(label, money_formula_or_blank, name_value=""):
            nonlocal rr
            _set_row_h(rr)

            c1 = ws.cell(rr, col_line, label)
            c1.alignment = center
            c1.border = border
            _set_base_font(c1, force_bold=True)

            c2 = ws.cell(rr, col_name, name_value or "")
            c2.border = box_border
            c2.alignment = left
            _set_base_font(c2)

            c3 = ws.cell(rr, col_money, money_formula_or_blank)
            c3.border = box_border
            c3.alignment = right
            c3.number_format = NUM_FMT_MONEY_HIDE0
            _set_base_font(c3)

            rr += 1

        if achievers:
            for lid in achievers:
                _write_row(str(lid), f"={int(default_money)}", names_for_line(str(lid), name_map))
        else:
            _write_row("（無）", "", "")
            return rr  # ✅ 直接結束，不加 QCA/QCB/補貨

        open_rng = f"$A$2:$A${last_row}"
        ach_rng = f"$L${first_list_row}:$L${rr-1}"

        open_qca = "(" + countif_sum(open_rng, QCA_LIST) + ")"
        ach_qca  = "(" + countif_sum(ach_rng,  QCA_LIST) + ")"
        open_qcb = "(" + countif_sum(open_rng, QCB_LIST) + ")"
        ach_qcb  = "(" + countif_sum(ach_rng,  QCB_LIST) + ")"

        open_all = f'COUNTIF({open_rng},"GT-*")'
        ach_all  = f'COUNTIF({ach_rng},"GT-*")'

        qca_formula = f'=IF({open_qca}=0,"",TRUNC({ach_qca}/{open_qca}*{qc_mult},0))'
        qcb_formula = f'=IF({open_qcb}=0,"",TRUNC({ach_qcb}/{open_qcb}*{qc_mult},0))'
        restock_formula = f'=IF({open_all}=0,"",TRUNC({ach_all}/{open_all}*{restock_mult},0))'

        _write_row("QCA", qca_formula, names_for_line("QCA", name_map))
        _write_row("QCB", qcb_formula, names_for_line("QCB", name_map))
        _write_row("補貨", restock_formula, names_for_line("補貨", name_map))

        return rr

    start = last_row + 3
    next_row = draw_achiever_block(
        start_row=start,
        title="上午達標名單（PCS≥上午應達成）",
        achievers=am_achievers,
        default_money=AM_DEFAULT_MONEY,
        qc_mult=50,
        restock_mult=100,
    )
    draw_achiever_block(
        start_row=next_row,
        title="下午達標名單（PCS≥下午目標）",
        achievers=pm_achievers,
        default_money=PM_DEFAULT_MONEY,
        qc_mult=100,
        restock_mult=50,
    )

    ws.freeze_panes = "A2"
    return ws


@st.cache_data(show_spinner=False)
def parse_source_file(raw: bytes, filename: str):
    df = robust_read_bytes(raw, filename)
    df, c_pickdate, c_packqty, c_cweight, c_lineid, c_stotype = normalize_columns(df)
    df2, line_base, split = build_hourly_metrics(df, c_pickdate, c_packqty, c_cweight, c_lineid, c_stotype)

    dates = sorted(df2["PICK_DATE"].dropna().unique())
    if not dates:
        raise ValueError("PICKDATE 解析後沒有日期資料，請確認 PICKDATE 欄位內容。")

    date_to_hours = {}
    date_to_lineids_all = {}
    for d in dates:
        hours = sorted(df2.loc[df2["PICK_DATE"] == d, "HOUR"].dropna().unique().tolist())
        date_to_hours[d] = [int(x) for x in hours]
        lineids = sorted(df2.loc[df2["PICK_DATE"] == d, c_lineid].dropna().astype(str).str.strip().unique().tolist())
        date_to_lineids_all[d] = lineids

    return df2, line_base, split, dates, date_to_hours, date_to_lineids_all, c_lineid, c_stotype


def excel_bytes_from_inputs(
    line_base, split,
    dates, date_to_hours, date_to_lineids_filtered,
    c_lineid, c_stotype,
    get_table_by_date_func,
    name_map
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    summary_inputs = {}

    for d in dates:
        hours = date_to_hours.get(d, [])
        lineids = date_to_lineids_filtered.get(d, [])
        if not hours or not lineids:
            continue

        mp_table = get_table_by_date_func(d)
        manpower_map = manpower_table_to_map(d, mp_table)

        sub_base = line_base[line_base["PICK_DATE"] == d]
        line_base_map = {}
        for lid in lineids:
            tmp = sub_base[sub_base[c_lineid].astype(str).str.strip() == str(lid)]
            line_base_map[(str(lid), "PCS")] = {int(r["HOUR"]): r["PCS"] for _, r in tmp.iterrows()}
            line_base_map[(str(lid), "加權PCS")] = {int(r["HOUR"]): r["加權PCS"] for _, r in tmp.iterrows()}

        sub_split = split[split["PICK_DATE"] == d]
        split_map = {}
        for lid in lineids:
            tmpL = sub_split[sub_split[c_lineid].astype(str).str.strip() == str(lid)]
            for t in ["GSO", "GXSO"]:
                tmpT = tmpL[tmpL[c_stotype].astype(str).str.strip() == t]
                split_map[(str(lid), t, "PCS")] = {int(r["HOUR"]): r["PCS"] for _, r in tmpT.iterrows()}
                split_map[(str(lid), t, "加權PCS")] = {int(r["HOUR"]): r["加權PCS"] for _, r in tmpT.iterrows()}

        date_ws = write_hourly_sheet(
            wb=wb,
            sheet_name=str(d),
            date_value=d,
            hours=hours,
            lineids=lineids,
            line_base_map=line_base_map,
            split_map=split_map,
            manpower_map=manpower_map,
        )

        am_ach, pm_ach = compute_achievers(d, lineids, hours, line_base_map, manpower_map)

        _, hour_to_col = find_hour_header_row(date_ws)
        line_map2 = parse_line_rows(date_ws)

        am_cols = [hour_to_col[h] for h in AM_HOURS if h in hour_to_col]
        pm_cols = [hour_to_col[h] for h in PM_HOURS if h in hour_to_col]

        summary_inputs[d] = (date_ws, line_map2, am_cols, pm_cols, am_ach, pm_ach)

    for d in dates:
        pack = summary_inputs.get(d)
        if not pack:
            continue
        date_ws, line_map2, am_cols, pm_cols, am_ach, pm_ach = pack
        if date_ws is None or not line_map2:
            continue

        build_summary_sheet_with_achievers(
            wb=wb,
            summary_name=f"彙總_{str(d)}",
            date_ws=date_ws,
            line_map=line_map2,
            am_cols=am_cols,
            pm_cols=pm_cols,
            am_achievers=am_ach,
            pm_achievers=pm_ach,
            name_map=name_map,
        )

    sum_sheets = [sn for sn in wb.sheetnames if sn.startswith("彙總_")]
    other_sheets = [sn for sn in wb.sheetnames if not sn.startswith("彙總_")]
    wb._sheets = [wb[sn] for sn in sum_sheets + other_sheets]

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# =========================
# UI
# =========================
st.set_page_config(page_title="大豐物流 - 出貨作業線產能", page_icon="📦", layout="wide")
inject_logistics_theme()
set_page("📦 出貨作業線產能（可手動輸入人力）")

card_open("上傳來源檔")
up = st.file_uploader(
    "請上傳來源檔案（含 PICKDATE / PACKQTY / Cweight / LINEID / STO_TYPE）",
    type=["xlsx", "xlsm", "xls", "csv"],
    accept_multiple_files=False
)
card_close()

if not up:
    st.info("請先上傳檔案。")
    st.stop()

raw = up.getvalue()
filename = up.name

# 檔案變更時清理舊狀態（避免混到上一份）
sig = hashlib.md5(raw).hexdigest()
if st.session_state.get("src_sig_v5") != sig:
    for k in list(st.session_state.keys()):
        if (
            k.startswith("mp_store_") or k.startswith("mp_rev_") or k.startswith("mp_schema_")
            or k.startswith("am_fill_") or k.startswith("pm_fill_")
            or k.startswith("excel_bytes_oneclick") or k.startswith("excel_name_oneclick")
        ):
            del st.session_state[k]
    st.session_state["src_sig_v5"] = sig

with st.spinner("解析檔案中..."):
    df2, line_base, split, dates, date_to_hours, date_to_lineids_all, c_lineid, c_stotype = parse_source_file(raw, filename)

st.success(f"讀取完成：共 {len(dates)} 天")

# 1) 不列入計算 LINE ID（手動輸入）
card_open("設定：不列入計算的 LINE ID（手動輸入）")
exclude_raw = st.text_input(
    "輸入要排除的 LINE ID（逗號/空白/換行分隔）",
    value="F0026, FUBOX, SORT",
    key="exclude_lineid_v5"
)
exclude_set = {x.strip() for x in re.split(r"[,\s]+", (exclude_raw or "").strip()) if x.strip()}
st.caption(f"目前排除：{', '.join(sorted(exclude_set)) if exclude_set else '（無）'}")
card_close()

# 2) 名單貼上
card_open("名單貼上：Line ID → 姓名（可自行貼上）")
st.caption("格式：每行第一欄是 Line ID，後面欄位是姓名（建議用 Tab 分隔）。")
name_tsv = st.text_area("貼上名單", value=st.session_state.get("name_tsv_v5", DEFAULT_NAME_TSV), height=220, key="name_tsv_v5")
name_map = parse_name_tsv(name_tsv)
st.caption(f"已載入名單：{len(name_map)} 個 Key")
card_close()

# 依排除清單產生每日期的 lineids（用於人力輸入/計算/匯出）
date_to_lineids = {}
for d in dates:
    all_ids = [str(x).strip() for x in date_to_lineids_all.get(d, []) if str(x).strip()]
    date_to_lineids[d] = [lid for lid in all_ids if lid not in exclude_set]

# 3) 人力輸入（rev 改 key 刷新 widget）
card_open("手動輸入人力（每小時 / 可小數）")
st.caption("✅ 已修正：不會再跳格回空白；快速填入/清空/排除變更都能立即刷新。")

tabs = st.tabs([str(d) for d in dates])

for i, d in enumerate(dates):
    with tabs[i]:
        hours = date_to_hours.get(d, [])
        lineids = date_to_lineids.get(d, [])

        if not hours or not lineids:
            st.warning("此日期沒有可用的 Hour/Line（可能都被排除）")
            continue

        store_key = f"mp_store_{str(d)}"
        rev_key = f"mp_rev_{str(d)}"
        schema_key = f"mp_schema_{str(d)}"

        if rev_key not in st.session_state:
            st.session_state[rev_key] = 0

        if store_key not in st.session_state:
            st.session_state[store_key] = build_default_manpower_table(lineids, hours)
        else:
            st.session_state[store_key] = reconcile_manpower_table(st.session_state[store_key], lineids, hours)

        schema_sig = (
            tuple(st.session_state[store_key]["Line ID"].astype(str).tolist()),
            tuple([c for c in st.session_state[store_key].columns if c != "Line ID"])
        )
        if st.session_state.get(schema_key) != schema_sig:
            st.session_state[schema_key] = schema_sig
            st.session_state[rev_key] += 1

        editor_key = f"mp_editor_{str(d)}_{st.session_state[rev_key]}"

        c1, c2, c3 = st.columns([1, 1, 2])
        with c1:
            am_fill = st.text_input(f"上午快速填入（{d}，套用 8-12）", value="", key=f"am_fill_{str(d)}")
        with c2:
            pm_fill = st.text_input(f"下午快速填入（{d}，套用 13-18）", value="", key=f"pm_fill_{str(d)}")
        with c3:
            st.caption("快速填入：輸入數字後按按鈕，可一鍵填滿該時段所有 Line。")

        def _apply_fill(target_hours, val: str):
            val = (val or "").strip()
            if val == "":
                return
            dfm = st.session_state[store_key].copy()
            for h in target_hours:
                hs = str(int(h))
                if hs in dfm.columns:
                    dfm.loc[:, hs] = val
            st.session_state[store_key] = dfm
            st.session_state[rev_key] += 1
            st.rerun()

        b1, b2, b3 = st.columns([1, 1, 1])
        with b1:
            if st.button("套用上午快速填入", key=f"btn_am_{str(d)}", use_container_width=True):
                _apply_fill(AM_HOURS, am_fill)
        with b2:
            if st.button("套用下午快速填入", key=f"btn_pm_{str(d)}", use_container_width=True):
                _apply_fill(PM_HOURS, pm_fill)
        with b3:
            if st.button("全部清空", key=f"btn_clear_{str(d)}", use_container_width=True):
                st.session_state[store_key] = build_default_manpower_table(lineids, hours)
                st.session_state[rev_key] += 1
                st.rerun()

        col_cfg = {"Line ID": st.column_config.TextColumn("Line ID", disabled=True)}
        for h in hours:
            col_cfg[str(int(h))] = st.column_config.TextColumn(str(int(h)))

        edited_df = st.data_editor(
            st.session_state[store_key],
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config=col_cfg,
            key=editor_key
        )
        st.session_state[store_key] = edited_df

card_close()

# 4) 即時達標狀況
st.markdown("---")
card_open("✅ Line ID 達標狀況（即時刷新）")

for d in dates:
    hours = date_to_hours.get(d, [])
    lineids = date_to_lineids.get(d, [])
    if not hours or not lineids:
        continue

    store_key = f"mp_store_{str(d)}"
    mp_table = st.session_state.get(store_key)
    mp_map = manpower_table_to_map(d, mp_table)

    lb_map = build_line_base_map_for_date(d, lineids, line_base, c_lineid)
    df_status = build_achv_status_table(d, lineids, hours, lb_map, mp_map, name_map)

    if df_status.empty:
        st.subheader(str(d))
        st.info("此日期沒有可顯示的達標資料。")
        continue

    am_ok = (df_status["上午達標"] == "✅").sum()
    pm_ok = (df_status["下午達標"] == "✅").sum()
    total = len(df_status)

    st.subheader(f"{d}（上午達標 {am_ok}/{total}｜下午達標 {pm_ok}/{total}）")
    st.dataframe(style_achv(df_status), use_container_width=True, hide_index=True)

card_close()

# =========================
# ✅ 真正：產出excel = 下載excel（1-click）
# =========================
st.markdown("---")
card_open("⬇️ 匯出 Excel（真正一鍵：產出=下載）")

base = os.path.splitext(os.path.basename(filename))[0]
out_name = st.text_input(
    "輸出檔名",
    value=f"{base}_每小時戰情表_含上午下午彙總.xlsx",
    key="out_name_v5"
)

def _get_table_by_date(d):
    return st.session_state.get(f"mp_store_{str(d)}")

if st.button("⬇️ 產出並下載 Excel（1-click）", type="primary", use_container_width=True):
    with st.spinner("Excel 產生中..."):
        xbytes = excel_bytes_from_inputs(
            line_base=line_base,
            split=split,
            dates=dates,
            date_to_hours=date_to_hours,
            date_to_lineids_filtered=date_to_lineids,
            c_lineid=c_lineid,
            c_stotype=c_stotype,
            get_table_by_date_func=_get_table_by_date,
            name_map=name_map,
        )

    st.session_state["excel_bytes_oneclick"] = xbytes
    st.session_state["excel_name_oneclick"] = out_name

    b64 = base64.b64encode(xbytes).decode("utf-8")
    safe_name = (out_name or "export.xlsx").replace('"', "").replace("\\", "_").replace("/", "_")

    components.html(
        f"""
        <html><body>
        <a id="dl" download="{safe_name}"
           href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}">
        </a>
        <script>
          document.getElementById('dl').click();
        </script>
        </body></html>
        """,
        height=0,
    )

    st.success("已觸發下載。若瀏覽器阻擋自動下載，請用下方「備援下載」按鈕。")

# 備援下載（瀏覽器策略阻擋自動下載時）
if st.session_state.get("excel_bytes_oneclick"):
    st.download_button(
        "⬇️ 備援下載（瀏覽器阻擋自動下載時使用）",
        data=st.session_state["excel_bytes_oneclick"],
        file_name=st.session_state.get("excel_name_oneclick", "export.xlsx"),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

card_close()
