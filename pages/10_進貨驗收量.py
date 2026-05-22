# pages/10_進貨驗收量.py
import io
import re
import pandas as pd
import streamlit as st

from common_ui import inject_logistics_theme, set_page, card_open, card_close

st.set_page_config(page_title="進貨驗收量｜大樹KPI", page_icon="📥", layout="wide")
inject_logistics_theme()

SHEET_DEFAULT = "採購單驗收量明細"

# ✅ 統一用「標準欄位名」做運算（先把報表欄位自動對照到這些名稱）
REQ_COLS = ["入庫類型", "驗收入庫數量", "供應商代號", "DC採購單號", "商品品號"]

# ✅ 你的檔案實際會出現的同義欄位（可再加）
COL_ALIASES = {
    "入庫類型": ["入庫類型", "入庫型態", "入庫类型"],
    "驗收入庫數量": ["驗收入庫數量", "驗收入庫量", "驗收入庫", "驗收入庫量", "驗收入庫數量"],
    "供應商代號": ["供應商代號", "廠商代號", "供應商編號", "廠商編號"],
    "DC採購單號": ["DC採購單號", "DC採購單号", "DC採購單", "採購單號(DC)"],
    # ⚠️ 若真的沒有 DC採購單號，才退回採購單號
    "__FALLBACK_DC_ORDER__": ["採購單號"],
    "商品品號": ["商品品號", "商品代號", "商品料號", "品號", "料號"],
}


def _norm_key(s: str) -> str:
    # 去空白、全形空白、常見符號
    s = str(s)
    s = s.replace("\u3000", " ")
    s = re.sub(r"\s+", "", s)
    return s.strip()


def _normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip().replace("\u3000", " ") for c in df.columns]
    return df


def _apply_column_aliases(df: pd.DataFrame) -> pd.DataFrame:
    """
    把「報表欄位」自動對照成「標準欄位」(REQ_COLS)
    - 只在標準欄位不存在時才會做 rename
    - DC採購單號：若沒有才用 採購單號 當 fallback
    """
    df = df.copy()
    cols = list(df.columns)

    # 建一個 normalized -> 原始欄位名 的 lookup
    norm_map = {}
    for c in cols:
        nk = _norm_key(c)
        if nk not in norm_map:
            norm_map[nk] = c

    def find_col(candidates):
        for cand in candidates:
            # 先直接命中
            if cand in df.columns:
                return cand
            # 再用 normalized 命中
            nk = _norm_key(cand)
            if nk in norm_map:
                return norm_map[nk]
        return None

    rename_map = {}

    # 一般欄位
    for target, candidates in COL_ALIASES.items():
        if target.startswith("__"):
            continue
        if target in df.columns:
            continue
        hit = find_col(candidates)
        if hit and hit != target:
            rename_map[hit] = target

    # DC採購單號 fallback（只有在真的沒有 DC採購單號 時才用「採購單號」）
    if "DC採購單號" not in df.columns:
        fb = find_col(COL_ALIASES["__FALLBACK_DC_ORDER__"])
        if fb:
            rename_map[fb] = "DC採購單號"

    if rename_map:
        df = df.rename(columns=rename_map)

    return df


def _is_empty_row(vals) -> bool:
    for v in vals:
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return False
    return True


def _trim_trailing_nones(vals):
    last = -1
    for i, v in enumerate(vals):
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        last = i
    if last < 0:
        return []
    return vals[: last + 1]


def _find_header_row(rows, required_cols, scan_rows=250):
    req = set([c.strip() for c in required_cols])
    for i, r in enumerate(rows[:scan_rows]):
        cand = [str(x).strip() if x is not None else "" for x in r]
        cand = [c for c in cand if c]
        cand_set = set(cand)
        if req.issubset(cand_set):
            return i
    return None


def _read_xlsb_with_pyxlsb(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    try:
        from pyxlsb import open_workbook
    except Exception as e:
        raise ImportError("讀取 .xlsb 需要安裝 pyxlsb（requirements.txt 加上 pyxlsb）。") from e

    bio = io.BytesIO(file_bytes)
    rows = []

    with open_workbook(bio) as wb:
        if sheet_name not in wb.sheets:
            raise KeyError(f"找不到工作表：{sheet_name}（目前工作表：{', '.join(wb.sheets)}）")

        with wb.get_sheet(sheet_name) as sh:
            for row in sh.rows():
                vals = [c.v for c in row]
                vals = _trim_trailing_nones(vals)
                rows.append(vals)

    if not rows:
        return pd.DataFrame()

    # 先用「標準欄位」找表頭；找不到就退回第一個非空列
    header_idx = _find_header_row(rows, list(set(sum([v for k, v in COL_ALIASES.items() if not k.startswith("__")], []))), scan_rows=250)
    if header_idx is None:
        header_idx = next((i for i, r in enumerate(rows[:250]) if not _is_empty_row(r)), None)

    if header_idx is None:
        return pd.DataFrame()

    header = [str(x).strip() if x is not None else "" for x in rows[header_idx]]
    header = [h if h else f"未命名欄位_{i+1}" for i, h in enumerate(header)]

    data_rows = rows[header_idx + 1 :]

    # 去掉前面空白列
    while data_rows and _is_empty_row(data_rows[0]):
        data_rows = data_rows[1:]

    cleaned = []
    empty_run = 0
    for r in data_rows:
        if _is_empty_row(r):
            empty_run += 1
            if empty_run >= 30:
                break
            continue
        empty_run = 0
        cleaned.append(r)

    if not cleaned:
        return pd.DataFrame(columns=header)

    max_len = len(header)
    fixed = []
    for r in cleaned:
        rr = list(r[:max_len]) + [None] * max(0, max_len - len(r))
        fixed.append(rr)

    df = pd.DataFrame(fixed, columns=header)
    df = _normalize_cols(df)
    df = _apply_column_aliases(df)
    return df


def _get_sheet_names(file_name: str, file_bytes: bytes):
    ext = file_name.lower().split(".")[-1]
    bio = io.BytesIO(file_bytes)

    if ext == "xlsb":
        from pyxlsb import open_workbook
        with open_workbook(bio) as wb:
            return list(wb.sheets)

    from openpyxl import load_workbook
    wb = load_workbook(bio, read_only=True, data_only=True)
    return wb.sheetnames


@st.cache_data(show_spinner=False)
def _read_excel_bytes(file_name: str, file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    ext = file_name.lower().split(".")[-1]
    bio = io.BytesIO(file_bytes)

    if ext == "xlsb":
        return _read_xlsb_with_pyxlsb(file_bytes, sheet_name)

    # xlsx / xlsm
    df = pd.read_excel(bio, sheet_name=sheet_name, engine="openpyxl")
    df = _normalize_cols(df)
    df = _apply_column_aliases(df)
    return df


def _compute_stats(df: pd.DataFrame, inbound_type: str) -> dict:
    df = df.copy()

    # 欄位保護
    for c in REQ_COLS:
        if c not in df.columns:
            raise KeyError(f"找不到必要欄位：{c}")

    df_type = df[df["入庫類型"].astype(str).str.strip().eq(inbound_type)].copy()
    df_type["驗收入庫數量"] = pd.to_numeric(df_type["驗收入庫數量"], errors="coerce").fillna(0)

    return {
        "type": inbound_type,
        "unique_suppliers": int(df_type["供應商代號"].nunique(dropna=True)),
        "unique_dc_orders": int(df_type["DC採購單號"].nunique(dropna=True)),
        "unique_products": int(df_type["商品品號"].nunique(dropna=True)),
        "total_qty": float(df_type["驗收入庫數量"].sum()),
    }


def _fmt_qty(x: float) -> str:
    if abs(x - round(x)) < 1e-9:
        return f"{int(round(x)):,}"
    return f"{x:,.2f}"


def main():
    set_page(
        "進貨驗收量",
        icon="📥",
        subtitle="大樹KPI｜每日驗收報表｜GPO / GXPO 統計",
    )

    card_open("📌 上傳檔案")
    up = st.file_uploader("上傳 .xlsb 或 .xlsx", type=["xlsb", "xlsx", "xlsm"])
    card_close()

    if not up:
        st.info("請先上傳檔案後再執行統計。")
        return

    file_bytes = up.getvalue()

    # 工作表清單
    try:
        sheet_names = _get_sheet_names(up.name, file_bytes)
    except Exception as e:
        st.error(f"讀取工作表清單失敗：{e}")
        st.stop()

    default_idx = sheet_names.index(SHEET_DEFAULT) if SHEET_DEFAULT in sheet_names else 0

    card_open("⚙️ 讀取設定")
    sheet_name = st.selectbox("工作表名稱", options=sheet_names, index=default_idx)
    card_close()

    with st.spinner("讀取資料中..."):
        try:
            df = _read_excel_bytes(up.name, file_bytes, sheet_name)
        except Exception as e:
            st.error(f"讀取失敗：{e}")
            st.stop()

    if df.empty:
        st.warning(
            "已成功讀取檔案，但這張工作表資料是空的或沒有可解析的值。\n\n"
            "若這份報表是公式/樞紐即時產生：建議先用 Excel 開啟一次 → 等計算完成 → 另存為 .xlsx 再上傳。"
        )
        st.write("目前欄位：", list(df.columns))
        return

    # ✅ 顯示「欄位對照後」的欄名，方便你確認
    with st.expander("🔎 檢視目前欄位（已自動對照）", expanded=False):
        st.write(list(df.columns))

    # 必要欄位檢查
    missing = [c for c in REQ_COLS if c not in df.columns]
    if missing:
        st.error(f"缺少必要欄位：{', '.join(missing)}")
        st.write("目前欄位：", list(df.columns))
        st.stop()

    # 統計
    types = ["GPO", "GXPO"]
    stats_rows = []
    for t in types:
        stats_rows.append(_compute_stats(df, t))

    overall_unique_suppliers = int(df["供應商代號"].nunique(dropna=True))

    card_open("📊 統計結果")
    cols = st.columns(len(types))
    for i, s in enumerate(stats_rows):
        with cols[i]:
            st.subheader(f"{s['type']} 類型")
            st.metric("不重複供應商代號", f"{s['unique_suppliers']:,} 筆")
            st.metric("不重複 DC採購單號", f"{s['unique_dc_orders']:,} 筆")
            st.metric("不重複 商品品號", f"{s['unique_products']:,} 筆")
            st.metric("驗收入庫數量總量", _fmt_qty(s["total_qty"]))

    st.divider()
    st.metric("總明細：不重複供應商代號總數", f"{overall_unique_suppliers:,} 筆")
    card_close()

    # 匯出
    out_df = pd.DataFrame(
        [
            {
                "入庫類型": r["type"],
                "不重複供應商代號數": r["unique_suppliers"],
                "不重複DC採購單號數": r["unique_dc_orders"],
                "不重複商品品號數": r["unique_products"],
                "驗收入庫數量總量": r["total_qty"],
            }
            for r in stats_rows
        ]
    )
    out_df.loc[len(out_df)] = {
        "入庫類型": "總明細",
        "不重複供應商代號數": overall_unique_suppliers,
        "不重複DC採購單號數": None,
        "不重複商品品號數": None,
        "驗收入庫數量總量": None,
    }

    card_open("📤 匯出")
    st.dataframe(out_df, use_container_width=True)

    csv_bytes = out_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
    st.download_button("下載 CSV", data=csv_bytes, file_name="進貨驗收量_統計.csv", mime="text/csv")
    card_close()


if __name__ == "__main__":
    main()
