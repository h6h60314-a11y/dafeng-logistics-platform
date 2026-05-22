# pages/27_QC未上架比對.py
# -*- coding: utf-8 -*-
import io
from collections import defaultdict
from datetime import datetime, date
from typing import Optional, Tuple, Dict

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import copy as _copy

from common_ui import inject_logistics_theme, set_page, card_open, card_close


# =============================
# 參數
# =============================
QC_KEY_HEADER = "商品"
UN_KEY_HEADER = "商品碼"
UN_DATE_HEADER = "進貨日"
UNIT_HEADER = "可移動單位"
BARCODE_HEADER = "國際條碼"
BATCH_HEADER = "批號"

MATCH_SHEET_NAME = "符合未上架明細"

# ✅ 你指定要刪除的欄位
DELETE_HEADERS = [
    "移動的數量",
    "目的儲位",
    "可移動單位至",
    "計量單位由",
    "到包裝碼",
    "已試算",
    "已揀取",
]

# ✅ 你指定的欄寬（A~K）
COLUMN_WIDTHS = {
    "A": 11,
    "B": 7.71,
    "C": 7,
    "D": 14.57,
    "E": 52.43,
    "F": 11.14,
    "G": 5.29,
    "H": 12,
    "I": 19.29,
    "J": 9.57,
    "K": 10.57,
}

# ✅ 你指定的列高（全部）
ROW_HEIGHT = 15.75


# =============================
# UI CSS（背景 + 上傳框風格）
# =============================
def _page_css():
    st.markdown(
        r"""
<style>
div[data-testid="stAppViewContainer"]{
  background: linear-gradient(
    180deg,
    rgba(232,245,255,1) 0%,
    rgba(244,250,255,1) 34%,
    rgba(255,255,255,1) 100%
  ) !important;
}

.qc-chips{
  margin-top: 4px;
  font-size: 12.5px;
  font-weight: 800;
  color: rgba(15,23,42,.62);
}
.qc-chips .sep{ margin: 0 8px; opacity:.55; }

.qc-u-label{
  font-size: 13.5px;
  font-weight: 900;
  color: rgba(15,23,42,.86);
  margin: 4px 0 6px 0;
}

section[data-testid="stFileUploadDropzone"]{
  border: 1px solid rgba(148,163,184,.35) !important;
  border-radius: 14px !important;
  background: rgba(255,255,255,1) !important;
  padding: 12px 14px !important;
}
section[data-testid="stFileUploadDropzone"]:hover{
  border-color: rgba(59,130,246,.35) !important;
  box-shadow: 0 6px 18px rgba(59,130,246,.08);
}

section[data-testid="stFileUploadDropzone"] button{
  border-radius: 10px !important;
  font-weight: 900 !important;
}

div[data-testid="stButton"] > button{
  border-radius: 12px !important;
  font-weight: 900 !important;
  padding: 9px 14px !important;
}

.qc-banner{
  background: rgba(219, 234, 254, .9);
  border: 1px solid rgba(59,130,246,.18);
  color: rgba(15,23,42,.86);
  border-radius: 10px;
  padding: 10px 12px;
  font-weight: 900;
  font-size: 13px;
  margin-top: 14px;
}
</style>
""",
        unsafe_allow_html=True,
    )


# =============================
# 工具：定位欄位 / 文字格式
# =============================
def get_ws(wb, sheet_name: Optional[str]):
    return wb[sheet_name] if sheet_name else wb.worksheets[0]


def find_header_col(ws, header_name: str, header_row: int = 1) -> Optional[int]:
    # exact
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip() == header_name:
            return c
    # contains
    target = header_name.strip()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and target in v.strip():
            return c
    return None


def zero_run_width(number_format: str) -> int:
    if not number_format:
        return 0
    fmt = number_format.split(";")[0]
    best = cur = 0
    for ch in fmt:
        if ch == "0":
            cur += 1
            best = max(best, cur)
        else:
            cur = 0
    return best


def normalize_code(value, fmt: str, fallback_width: int = 0) -> str:
    """只用於比對（不回寫）"""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")

    width = max(zero_run_width(fmt or ""), fallback_width)

    if isinstance(value, int):
        s = str(value)
        return s.zfill(width) if width >= 2 else s

    if isinstance(value, float):
        if abs(value - round(value)) < 1e-9:
            iv = int(round(value))
            s = str(iv)
            return s.zfill(width) if width >= 2 else s
        return str(value)

    return str(value).strip()


def normalize_unit(value) -> str:
    """只用於比對（不回寫）"""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _infer_digit_width(ws, col_idx: int, scan_limit: int = 50000) -> int:
    """推斷碼長（未上架明細）"""
    if col_idx is None:
        return 0
    w = 0
    end_r = min(ws.max_row, scan_limit)
    for r in range(2, end_r + 1):
        cell = ws.cell(row=r, column=col_idx)
        v = cell.value
        if v is None:
            continue

        fmt = getattr(cell, "number_format", "") or ""
        w = max(w, zero_run_width(fmt))

        if isinstance(v, str):
            s = v.strip()
            if s.isdigit():
                w = max(w, len(s))
    return w


def _pad_digits_for_compare(s: str, width: int) -> str:
    """只拿來比對，不回寫到 QC"""
    if not s:
        return ""
    if width >= 2 and s.isdigit():
        return s.zfill(width)
    return s


def format_date_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return ""
    try:
        dtv = pd.to_datetime(s)
        return pd.Timestamp(dtv).strftime("%Y-%m-%d")
    except Exception:
        return s


# =============================
# ✅ 輸出格式：欄寬 + 列高（全部 15.75）
# =============================
def apply_qc_layout(ws):
    # 欄寬 A~K
    for col_letter, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = w

    # 列高：全部 15.75
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT


# =============================
# ✅ 輸出用：可移動單位補滿 10 碼（不足補0）+ 設文字
# =============================
def pad_unit_to_10(ws, unit_col: int, start_row: int = 2, width: int = 10):
    if not unit_col:
        return
    for r in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=r, column=unit_col)
        v = cell.value
        if v is None:
            continue

        if isinstance(v, float) and abs(v - round(v)) < 1e-9:
            s = str(int(round(v)))
        else:
            s = str(v).strip()

        if not s:
            continue
        if s.isdigit():
            s = s.zfill(width)

        cell.value = s
        cell.number_format = "@"
        if cell.alignment is None:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# =============================
# ✅ 輸出用：把指定欄位強制轉文字（避免科學記號/掉0）
# - barcode 不補0，只轉成字串
# - batch 依推寬度補0
# =============================
def force_text_digits(ws, col_idx: int, pad_width: int = 0, start_row: int = 2):
    if not col_idx:
        return
    for r in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        v = cell.value
        if v is None:
            continue

        if isinstance(v, float) and abs(v - round(v)) < 1e-9:
            s = str(int(round(v)))
        elif isinstance(v, int):
            s = str(v)
        else:
            s = str(v).strip()

        if not s:
            continue

        if pad_width >= 2 and s.isdigit():
            s = s.zfill(pad_width)

        cell.value = s
        cell.number_format = "@"
        if cell.alignment is None:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def infer_col_digit_width(ws, col_idx: int, header_row: int = 1, scan_limit: int = 50000) -> int:
    if not col_idx:
        return 0
    w = 0
    end_r = min(ws.max_row, scan_limit)
    for r in range(header_row + 1, end_r + 1):
        cell = ws.cell(row=r, column=col_idx)
        v = cell.value
        if v is None:
            continue
        fmt = getattr(cell, "number_format", "") or ""
        w = max(w, zero_run_width(fmt))
        if isinstance(v, str):
            s = v.strip()
            if s.isdigit():
                w = max(w, len(s))
    return w


# =============================
# ✅ 刪除指定欄位（輸出檔用）
# =============================
def delete_columns_by_headers(ws, headers_to_delete, header_row: int = 1):
    if not headers_to_delete:
        return

    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        headers.append("" if v is None else str(v).strip())

    to_delete_cols = set()
    for target in headers_to_delete:
        t = str(target).strip()
        if not t:
            continue
        for idx, h in enumerate(headers, start=1):
            if not h:
                continue
            if h == t or (t in h):
                to_delete_cols.add(idx)

    for col_idx in sorted(to_delete_cols, reverse=True):
        ws.delete_cols(col_idx, 1)

    try:
        if getattr(ws, "auto_filter", None) and ws.auto_filter.ref:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    except Exception:
        pass


def delete_cols_after_header(ws, header_name: str, header_row: int = 1):
    col = find_header_col(ws, header_name, header_row)
    if col is None:
        return
    if ws.max_column > col:
        ws.delete_cols(col + 1, ws.max_column - col)


# =============================
# 轉換：DataFrames -> openpyxl Workbook（xls/xlsb 用）
# =============================
def _dfs_to_workbook(sheets: Dict[str, pd.DataFrame]) -> Workbook:
    wb = Workbook()
    if wb.worksheets:
        wb.remove(wb.worksheets[0])

    for sheet_name, df in sheets.items():
        name = str(sheet_name)[:31] if sheet_name else "Sheet1"
        ws = wb.create_sheet(title=name)

        headers = [("" if c is None else str(c)) for c in df.columns.tolist()]
        ws.append(headers)

        for row in df.itertuples(index=False, name=None):
            out_row = []
            for v in row:
                if v is None:
                    out_row.append(None)
                elif isinstance(v, float) and pd.isna(v):
                    out_row.append(None)
                else:
                    out_row.append(v)
            ws.append(out_row)

    return wb


# =============================
# 讀取：支援 xlsx/xlsm/xls/xlsb（回傳 mode）
# =============================
def _load_wb_from_upload(uploaded_file) -> Tuple[str, Workbook, str]:
    name = uploaded_file.name
    ext = (name.split(".")[-1] or "").lower()
    raw = uploaded_file.getvalue()
    bio = io.BytesIO(raw)

    if ext in ("xlsx", "xlsm"):
        keep_vba = (ext == "xlsm")
        wb = load_workbook(bio, keep_vba=keep_vba)
        return name, wb, "native"

    if ext == "xlsb":
        try:
            sheets = pd.read_excel(
                io.BytesIO(raw),
                engine="pyxlsb",
                sheet_name=None,
                dtype=str,
                keep_default_na=False,
            )
        except Exception as e:
            raise ValueError(f"讀取 .xlsb 失敗：{e}\n請確認 requirements.txt 有 pyxlsb")
        wb = _dfs_to_workbook(sheets)
        return name, wb, "converted"

    if ext == "xls":
        try:
            sheets = pd.read_excel(
                io.BytesIO(raw),
                engine="xlrd",
                sheet_name=None,
                dtype=str,
                keep_default_na=False,
            )
        except ModuleNotFoundError:
            raise ValueError(
                "目前環境缺少 xlrd，無法讀取 .xls。\n"
                "請在 requirements.txt 加上：xlrd==2.0.1\n"
                "或先用 Excel 另存為 .xlsx 再上傳。"
            )
        except Exception as e:
            raise ValueError(f"讀取 .xls 失敗：{e}\n建議先用 Excel 另存 .xlsx 再上傳。")
        wb = _dfs_to_workbook(sheets)
        return name, wb, "converted"

    raise ValueError(f"不支援的檔案格式：.{ext}\n支援：.xlsx / .xlsm / .xls / .xlsb")


# =============================
# 小工具：刪除非符合列（保留 header）
# =============================
def _delete_non_matched_rows(ws, keep_rows, header_rows: int = 1):
    max_row = ws.max_row
    start = header_rows + 1
    keep_set = set(r for r in keep_rows if start <= r <= max_row)
    keep_sorted = sorted(keep_set)

    segs = []
    cur = start
    for k in keep_sorted:
        if cur < k:
            segs.append((cur, k - 1))
        cur = k + 1
    if cur <= max_row:
        segs.append((cur, max_row))

    for s, e in reversed(segs):
        ws.delete_rows(s, e - s + 1)

    try:
        if getattr(ws, "auto_filter", None) and ws.auto_filter.ref:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    except Exception:
        pass


# =============================
# 主流程
# =============================
def process_wb(
    qc_wb,
    un_wb,
    qc_sheet_name: Optional[str] = None,
    un_sheet_name: Optional[str] = None,
) -> Tuple[int, bytes]:
    qc_ws = get_ws(qc_wb, qc_sheet_name)
    un_ws = get_ws(un_wb, un_sheet_name)

    qc_key_col = find_header_col(qc_ws, QC_KEY_HEADER, 1)
    qc_unit_col = find_header_col(qc_ws, UNIT_HEADER, 1)
    qc_barcode_col = find_header_col(qc_ws, BARCODE_HEADER, 1)
    qc_batch_col = find_header_col(qc_ws, BATCH_HEADER, 1)

    un_key_col = find_header_col(un_ws, UN_KEY_HEADER, 1)
    un_unit_col = find_header_col(un_ws, UNIT_HEADER, 1)
    un_date_col = find_header_col(un_ws, UN_DATE_HEADER, 1)

    if qc_key_col is None:
        raise ValueError(f"QC 找不到欄位：{QC_KEY_HEADER}")
    if qc_unit_col is None:
        raise ValueError(f"QC 找不到欄位：{UNIT_HEADER}")

    if un_key_col is None:
        raise ValueError(f"未上架明細找不到欄位：{UN_KEY_HEADER}")
    if un_unit_col is None:
        raise ValueError(f"未上架明細找不到欄位：{UNIT_HEADER}")
    if un_date_col is None:
        raise ValueError(f"未上架明細找不到欄位：{UN_DATE_HEADER}")

    # 1) 推估商品碼長（比對用）
    code_len = 0
    for r in range(2, un_ws.max_row + 1):
        cell = un_ws.cell(row=r, column=un_key_col)
        if isinstance(cell.value, str):
            s = cell.value.strip()
            if s.isdigit():
                code_len = max(code_len, len(s))
        else:
            code_len = max(code_len, zero_run_width(getattr(cell, "number_format", "") or ""))
    fallback_width = code_len or 6

    # 2) 推估可移動單位碼長（比對用）
    unit_width = _infer_digit_width(un_ws, un_unit_col)

    # 3) 建索引：(商品碼, 可移動單位) -> 進貨日
    date_sets = defaultdict(set)
    for r in range(2, un_ws.max_row + 1):
        code_cell = un_ws.cell(row=r, column=un_key_col)
        code = normalize_code(code_cell.value, getattr(code_cell, "number_format", ""), fallback_width)
        if code and code.isdigit():
            code = code.zfill(fallback_width)

        unit_cell = un_ws.cell(row=r, column=un_unit_col)
        unit = normalize_unit(unit_cell.value)
        unit = _pad_digits_for_compare(unit, unit_width)

        d_cell = un_ws.cell(row=r, column=un_date_col)
        d_str = format_date_value(d_cell.value)

        if code and unit and d_str:
            date_sets[(code, unit)].add(d_str)

    date_map: Dict[Tuple[str, str], str] = {k: "、".join(sorted(v)) for k, v in date_sets.items()}

    # 4) 新增/定位「進貨日」
    qc_date_col = find_header_col(qc_ws, "進貨日", 1)
    if qc_date_col is None:
        qc_date_col = qc_ws.max_column + 1
        hdr = qc_ws.cell(row=1, column=qc_date_col, value="進貨日")
        src_hdr = qc_ws.cell(row=1, column=qc_unit_col)
        try:
            hdr._style = _copy.copy(src_hdr._style)
        except Exception:
            pass
        hdr.alignment = _copy.copy(getattr(src_hdr, "alignment", Alignment(horizontal="center", vertical="center")))

    # 5) 填入進貨日 + 收集 match rows
    match_rows = []
    for r in range(2, qc_ws.max_row + 1):
        code_cell = qc_ws.cell(row=r, column=qc_key_col)
        code = normalize_code(code_cell.value, getattr(code_cell, "number_format", ""), fallback_width)
        if code and code.isdigit():
            code = code.zfill(fallback_width)

        unit_cell = qc_ws.cell(row=r, column=qc_unit_col)
        unit = normalize_unit(unit_cell.value)
        unit = _pad_digits_for_compare(unit, unit_width)

        d_str = date_map.get((code, unit), "")

        out_cell = qc_ws.cell(row=r, column=qc_date_col)
        out_cell.value = d_str
        out_cell.number_format = "@"
        out_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if d_str:
            match_rows.append(r)

    # ✅ 5.5) 可移動單位：補滿10碼
    pad_unit_to_10(qc_ws, qc_unit_col, start_row=2, width=10)

    # ✅ 5.6) 國際條碼：文字化（避免 E+12）
    if qc_barcode_col is not None:
        force_text_digits(qc_ws, qc_barcode_col, pad_width=0)

    # ✅ 5.7) 批號：推寬度後補0 + 文字化
    if qc_batch_col is not None:
        batch_w = infer_col_digit_width(qc_ws, qc_batch_col)
        force_text_digits(qc_ws, qc_batch_col, pad_width=batch_w)

    # ✅ 5.8) 刪除你指定的欄位
    delete_columns_by_headers(qc_ws, DELETE_HEADERS, header_row=1)

    # ✅ 5.9) 只保留到「進貨日」為止（右邊全部刪掉）
    delete_cols_after_header(qc_ws, "進貨日", header_row=1)

    # ✅ 5.10) 套用欄寬 + 列高（全部 15.75）
    apply_qc_layout(qc_ws)

    # 6) 產生符合分頁：複製 QC → 刪除不符合列
    if MATCH_SHEET_NAME in qc_wb.sheetnames:
        del qc_wb[MATCH_SHEET_NAME]

    mws = qc_wb.copy_worksheet(qc_ws)
    mws.title = MATCH_SHEET_NAME
    _delete_non_matched_rows(mws, keep_rows=match_rows, header_rows=1)

    # ✅ 符合分頁也套同樣欄寬/列高
    apply_qc_layout(mws)

    out = io.BytesIO()
    qc_wb.save(out)
    out.seek(0)
    return len(match_rows), out.getvalue()


# =============================
# Streamlit UI
# =============================
st.set_page_config(page_title="大豐物流 - 進貨課｜QC未上架比對", page_icon="🧾", layout="wide")
inject_logistics_theme()
_page_css()

set_page(
    "QC 未上架比對",
    icon="🧾",
    subtitle="輸出：可移動單位補10碼、刪指定欄位、只保留到進貨日；欄寬A~K固定、列高全 15.75。",
)

st.markdown(
    '<div class="qc-chips">雙條件比對<span class="sep">｜</span>可移動單位補10碼<span class="sep">｜</span>欄寬固定A~K<span class="sep">｜</span>列高全15.75</div>',
    unsafe_allow_html=True,
)

card_open("📁 檔案上傳")

st.markdown('<div class="qc-u-label">QC 明細（支援：.xlsx / .xlsm / .xls / .xlsb）</div>', unsafe_allow_html=True)
qc_file = st.file_uploader(
    "QC 明細",
    type=["xlsx", "xlsm", "xls", "xlsb"],
    accept_multiple_files=False,
    label_visibility="collapsed",
    key="qc_file",
)

st.markdown('<div style="height:10px"></div>', unsafe_allow_html=True)

st.markdown('<div class="qc-u-label">未上架明細（支援：.xlsx / .xlsm / .xls / .xlsb）</div>', unsafe_allow_html=True)
un_file = st.file_uploader(
    "未上架明細",
    type=["xlsx", "xlsm", "xls", "xlsb"],
    accept_multiple_files=False,
    label_visibility="collapsed",
    key="un_file",
)

qc_sheet_name = None
un_sheet_name = None

with st.expander("進階設定（工作表選擇）", expanded=False):
    c1, c2 = st.columns(2)
    with c1:
        if qc_file:
            try:
                _, qc_wb_preview, _ = _load_wb_from_upload(qc_file)
                qc_sheet_name = st.selectbox("QC 工作表", options=qc_wb_preview.sheetnames, index=0)
            except Exception as e:
                st.error(str(e))
    with c2:
        if un_file:
            try:
                _, un_wb_preview, _ = _load_wb_from_upload(un_file)
                un_sheet_name = st.selectbox("未上架明細 工作表", options=un_wb_preview.sheetnames, index=0)
            except Exception as e:
                st.error(str(e))

ready = bool(qc_file and un_file)
run = st.button("🚀 產出比對", disabled=not ready)

card_close()

status_msg = "請依序上傳：QC 明細 + 未上架明細"
xlsx_bytes = None
matched = None

if ready:
    status_msg = "檔案已就緒，可按「產出比對」"

if run:
    try:
        with st.spinner("處理中…"):
            _, qc_wb, _ = _load_wb_from_upload(qc_file)
            _, un_wb, _ = _load_wb_from_upload(un_file)

            matched, xlsx_bytes = process_wb(
                qc_wb=qc_wb,
                un_wb=un_wb,
                qc_sheet_name=qc_sheet_name,
                un_sheet_name=un_sheet_name,
            )
    except Exception as e:
        st.error(f"❌ 執行失敗：{e}")

if xlsx_bytes is not None:
    card_open("✅ 產出結果")
    st.success(f"完成！符合筆數：{matched}")
    out_name = f"QC未上架比對_輸出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.download_button(
        label="📥 下載輸出 Excel",
        data=xlsx_bytes,
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    card_close()

st.markdown(f'<div class="qc-banner">{status_msg}</div>', unsafe_allow_html=True)
