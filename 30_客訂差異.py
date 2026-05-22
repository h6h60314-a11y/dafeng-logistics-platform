import io
import re
import datetime as dt
from typing import Dict, List, Tuple, Optional, Any

import pandas as pd
import streamlit as st

from common_ui import (
    inject_logistics_theme,
    set_page,
    KPI,
    render_kpis,
    bar_topN,
    card_open,
    card_close,
    download_excel_card,
    sidebar_controls,
)

# =========================================================
# 參數
# =========================================================
TO_EXCLUDE_KEYWORDS = ["CGS", "JCPL", "QC99", "GREAT0001X", "GX010", "PD99"]
TO_EXCLUDE_PATTERN = re.compile("|".join(re.escape(k) for k in TO_EXCLUDE_KEYWORDS), flags=re.IGNORECASE)

INPUT_USER_CANDIDATES = ["記錄輸入人", "記錄輸入者", "建立人", "輸入人"]
REV_DT_CANDIDATES = ["修訂日期", "修訂時間", "修訂日", "異動時間", "修改時間"]

TARGET_EFF_DEFAULT = 20
IDLE_MIN_THRESHOLD_DEFAULT = 10

AM_START, AM_END = dt.time(7, 0, 0), dt.time(12, 30, 0)
PM_START, PM_END = dt.time(13, 30, 0), dt.time(23, 59, 59)

# ✅ 儲位類型（區碼3 → 類型）
STORAGE_TYPE_ZONES = {
    "輕型料架": ["001", "002", "003", "017", "016"],
    "落地儲": ["014", "018", "019", "020", "010", "081", "401", "402", "403", "015"],
    "重型低空": ["011", "012", "013", "031", "032", "033", "034", "035", "036", "037", "038"],
    "高空儲": [
        "021", "022", "023",
        "041", "042", "043",
        "051", "052", "053", "054", "055", "056", "057",
        "301", "302", "303", "304", "305", "306",
    ],
}
ZONE3_TO_STORAGE_TYPE = {z: t for t, zones in STORAGE_TYPE_ZONES.items() for z in zones}

# ✅ 既有代碼→姓名（仍保留）
NAME_MAP = {
    "20200924001": "黃雅君", "20210805001": "郭中合", "20220505002": "阮文青明",
    "20221221001": "阮文全", "20221222005": "謝忠龍", "20230119001": "陶春青",
    "20240926001": "陳莉娜", "20241011002": "林雙慧", "20250502001": "吳詩敏",
    "20250617001": "阮文譚", "20250617003": "喬家寶", "20250901009": "張寶萱",
    "G01": "0", "20201109003": "吳振凱", "09963": "黃謙凱",
    "20240313003": "阮曰忠", "20201109001": "梁冠如", "10003": "李茂銓",
    "20200922002": "葉欲弘", "20250923019": "阮氏紅深", "9963": "黃謙凱",
    "11399": "陳哲沅",
}

BREAK_RULES = [
    (dt.time(20, 45, 0), dt.time(22, 30, 0), 0,  "首≥20:45 且 末≤22:30 → 0 分鐘"),
    (dt.time(18, 30, 0), dt.time(20, 30, 0), 0,  "首≥18:30 且 末≤20:30 → 0 分鐘"),
    (dt.time(15, 30, 0), dt.time(18,  0, 0), 0,  "首≥15:30 且 末≤18:00 → 0 分鐘"),
    (dt.time(13, 30, 0), dt.time(15, 35, 0), 0,  "首≥13:30 且 末≤15:35 → 0 分鐘"),
    (dt.time(20, 45, 0), dt.time(23,  0, 0), 0,  "首≥20:45 且 末≤23:00 → 0 分鐘"),
    (dt.time(20,  0, 0), dt.time(22,  0, 0), 15, "首≥20:00 且 末≤22:00 → 15 分鐘"),
    (dt.time(18, 30, 0), dt.time(22,  0, 0), 15, "首≥18:30 且 末≤22:00 → 15 分鐘"),
    (dt.time(19,  0, 0), dt.time(22, 30, 0), 15, "首≥19:00 且 末≤22:30 → 15 分鐘"),
    (dt.time(13, 30, 0), dt.time(18,  0, 0), 15, "首≥13:30 且 末≤18:00 → 15 分鐘"),
    (dt.time(16,  0, 0), dt.time(20, 40, 0), 30, "首≥16:00 且 末≤20:40 → 30 分鐘"),
    (dt.time(15, 30, 0), dt.time(20, 30, 0), 30, "首≥15:30 且 末≤20:30 → 30 分鐘"),
    (dt.time(17,  0, 0), dt.time(22, 30, 0), 45, "首≥17:00 且 末≤22:30 → 45 分鐘"),
    (dt.time(15, 45, 0), dt.time(22, 30, 0), 45, "首≥15:45 且 末≤22:30 → 45 分鐘"),
    (dt.time(13, 30, 0), dt.time(20, 29, 0), 45, "首≥13:30 且 末≤20:29 → 45 分鐘"),
    (dt.time(13, 30, 0), dt.time(23,  0, 0), 60, "首≥13:30 且 末≤23:00 → 60 分鐘"),
    (dt.time(11,  0, 0), dt.time(17,  0, 0), 75, "首≥11:00 且 末≤17:00 → 75 分鐘"),
    (dt.time( 8,  0, 0), dt.time(17,  0, 0), 90, "首≥08:00 且 末≤17:00 → 90 分鐘"),
    (dt.time(10, 50, 0), dt.time(23,  0, 0), 120,"首≥10:50 且 末≤23:00 → 120 分鐘"),
    (dt.time( 8,  0, 0), dt.time(23,  0, 0), 135,"首≥08:00 且 末≤23:00 → 135 分鐘"),
]

# ✅ 預設排除空窗時段（可被 sidebar 覆蓋）
EXCLUDE_IDLE_RANGES_DEFAULT = [
    (dt.time(10,  0, 0), dt.time(10, 15, 0)),
    (dt.time(12, 30, 0), dt.time(13, 30, 0)),
    (dt.time(15, 30, 0), dt.time(15, 45, 0)),
    (dt.time(18,  0, 0), dt.time(18, 30, 0)),
    (dt.time(20, 30, 0), dt.time(20, 45, 0)),
]

# =========================================================
# 通用 helpers
# =========================================================
def _parse_time_any(x: Any) -> Optional[dt.time]:
    if x is None:
        return None
    if isinstance(x, dt.time):
        return x
    s = str(x).strip()
    if not s:
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$", s)
    if not m:
        return None
    hh = int(m.group(1))
    mm = int(m.group(2))
    ss = int(m.group(3) or 0)
    if not (0 <= hh <= 23 and 0 <= mm <= 59 and 0 <= ss <= 59):
        return None
    return dt.time(hh, mm, ss)


def _extract_zone3(s: Any) -> str:
    """從字串抓第一個 3 碼數字（001/014/301...）"""
    if s is None:
        return ""
    txt = str(s).strip()
    if not txt:
        return ""
    m = re.search(r"(\d{3})", txt)
    return m.group(1) if m else ""


def _map_storage_type(zone3: str) -> str:
    return ZONE3_TO_STORAGE_TYPE.get(str(zone3).strip(), "")


# =========================================================
# 上架人設定（session_state）
# =========================================================
PUTAWAY_PEOPLE_STATE_KEY = "putaway_people_settings"  # code -> {name, area}

def _get_putaway_people_settings() -> Dict[str, Dict[str, str]]:
    if PUTAWAY_PEOPLE_STATE_KEY not in st.session_state:
        st.session_state[PUTAWAY_PEOPLE_STATE_KEY] = {}
    return st.session_state[PUTAWAY_PEOPLE_STATE_KEY]

def _normalize_code(x: Any) -> str:
    return str(x).strip()

def render_putaway_people_settings_panel():
    settings = _get_putaway_people_settings()

    with st.sidebar.expander("📦 上架人設定（可中文姓名）", expanded=False):
        code = st.text_input("上架人代碼（可貼上）", key="putaway_person_code")
        name = st.text_input("姓名（中文可輸入）", key="putaway_person_name")
        area = st.selectbox("區域", ["低空", "高空"], index=0, key="putaway_person_area")

        c1, c2 = st.columns(2)
        with c1:
            if st.button("➕ 新增 / 更新", key="putaway_person_upsert"):
                c = _normalize_code(code)
                if not c:
                    st.error("請先輸入上架人代碼")
                else:
                    settings[c] = {"name": str(name).strip(), "area": str(area).strip()}
                    st.success(f"已更新：{c}")

        with c2:
            del_code = st.selectbox("刪除代碼", [""] + sorted(list(settings.keys())), key="putaway_person_del_code")
            if st.button("🗑️ 刪除", key="putaway_person_delete", disabled=(del_code == "")):
                settings.pop(del_code, None)
                st.success(f"已刪除：{del_code}")

        if settings:
            df = pd.DataFrame(
                [{"代碼": k, "姓名": v.get("name", ""), "區域": v.get("area", "")} for k, v in settings.items()]
            ).sort_values(["區域", "代碼"], ascending=[True, True])
            st.dataframe(df, use_container_width=True, hide_index=True)
        else:
            st.caption("（尚未設定）")


# =========================================================
# sidebar_controls 排除區間解析
# =========================================================
def _parse_exclude_windows(val: Any) -> List[Tuple[dt.time, dt.time]]:
    if val is None:
        return EXCLUDE_IDLE_RANGES_DEFAULT

    if isinstance(val, dict):
        for k in ("exclude_windows", "exclude_windows_times", "windows", "ranges", "exclude_ranges"):
            if k in val:
                return _parse_exclude_windows(val.get(k))
        return EXCLUDE_IDLE_RANGES_DEFAULT

    if isinstance(val, str):
        raw = val.strip()
        if not raw:
            return EXCLUDE_IDLE_RANGES_DEFAULT
        parts = re.split(r"[，,;；\n]+", raw)
        items = []
        for p in parts:
            p = p.strip()
            if not p:
                continue
            m = re.match(r"^(\d{1,2}:\d{2}(?::\d{2})?)\s*[-~～]\s*(\d{1,2}:\d{2}(?::\d{2})?)$", p)
            if m:
                items.append((m.group(1), m.group(2)))
        return _parse_exclude_windows(items) if items else EXCLUDE_IDLE_RANGES_DEFAULT

    if not isinstance(val, (list, tuple)):
        return EXCLUDE_IDLE_RANGES_DEFAULT

    out: List[Tuple[dt.time, dt.time]] = []
    for item in val:
        if isinstance(item, dict):
            s = _parse_time_any(item.get("start") or item.get("s") or item.get("from"))
            e = _parse_time_any(item.get("end") or item.get("e") or item.get("to"))
        elif isinstance(item, (list, tuple)) and len(item) >= 2:
            s = _parse_time_any(item[0])
            e = _parse_time_any(item[1])
        else:
            s, e = None, None

        if s and e and (dt.datetime.combine(dt.date.today(), s) < dt.datetime.combine(dt.date.today(), e)):
            out.append((s, e))

    return out if out else EXCLUDE_IDLE_RANGES_DEFAULT


def _extract_exclude_value_from_controls(controls: Dict[str, Any]) -> Any:
    if not isinstance(controls, dict) or not controls:
        return None
    for k in (
        "exclude_windows",
        "exclude_windows_times",
        "exclude_ranges",
        "exclude_idle_ranges",
        "idle_exclude_windows",
        "idle_exclude_ranges",
    ):
        if k in controls and controls.get(k):
            return controls.get(k)

    for k, v in controls.items():
        lk = str(k).lower()
        if ("exclude" in lk) and (("window" in lk) or ("range" in lk)) and v:
            return v
    return None


# =========================================================
# 讀檔（bytes）
# =========================================================
def read_excel_any_quiet_bytes(name: str, content: bytes) -> Dict[str, pd.DataFrame]:
    ext = (name.split(".")[-1] or "").lower()
    if ext in ("xlsx", "xlsm"):
        xl = pd.ExcelFile(io.BytesIO(content), engine="openpyxl")
        return {sn: pd.read_excel(xl, sheet_name=sn) for sn in xl.sheet_names}
    if ext == "xls":
        xl = pd.ExcelFile(io.BytesIO(content), engine="xlrd")
        return {sn: pd.read_excel(xl, sheet_name=sn) for sn in xl.sheet_names}
    if ext == "csv":
        for enc in ("utf-8-sig", "cp950", "big5"):
            try:
                return {"CSV": pd.read_csv(io.BytesIO(content), encoding=enc)}
            except Exception:
                continue
        raise Exception("CSV 讀取失敗（請確認編碼）")
    raise Exception("不支援的副檔名（僅支援 xlsx/xlsm/xls/csv）")


def _strip_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def find_first_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = [str(c).strip() for c in df.columns]
    s = set(cols)
    for name in candidates:
        if name in s:
            return name
    norm_map = {re.sub(r"[（）\(\)\s]", "", c): c for c in cols}
    for name in candidates:
        key = re.sub(r"[（）\(\)\s]", "", name)
        if key in norm_map:
            return norm_map[key]
    return None


def normalize_to_qc(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.upper().eq("QC")


def to_not_excluded_mask(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    return ~s.str.contains(TO_EXCLUDE_PATTERN, na=False)


def prepare_filtered_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    df = _strip_cols(df)
    if "由" not in df.columns or "到" not in df.columns:
        return pd.DataFrame()
    return df[normalize_to_qc(df["由"]) & to_not_excluded_mask(df["到"])].copy()


# =========================================================
# ✅ 棚別主檔：儲位→棚別
# =========================================================
def load_slot_master_bytes(upload_name: str, content: bytes) -> pd.DataFrame:
    sheets = read_excel_any_quiet_bytes(upload_name, content)
    df0 = None
    for _, d in sheets.items():
        if d is not None and not d.empty:
            df0 = d
            break
    if df0 is None or df0.empty:
        return pd.DataFrame()

    df0 = _strip_cols(df0)
    if "儲位" not in df0.columns or "棚別" not in df0.columns:
        return pd.DataFrame()

    out = df0[["儲位", "棚別"]].copy()
    out["儲位"] = out["儲位"].astype(str).str.strip()
    out["棚別"] = out["棚別"].astype(str).str.strip()
    out = out[out["儲位"].astype(str).str.len() > 0].copy()
    out = out.drop_duplicates(subset=["儲位"], keep="last")
    return out


# =========================================================
# 計算：休息 / 空窗 / clamp + 棚別比對筆數
# =========================================================
def break_minutes_for_span(first_dt: pd.Timestamp, last_dt: pd.Timestamp) -> Tuple[int, str]:
    if pd.isna(first_dt) or pd.isna(last_dt):
        return 0, "無時間資料"
    stt, edt = first_dt.time(), last_dt.time()
    for st_ge, ed_le, mins, tag in BREAK_RULES:
        if (stt >= st_ge) and (edt <= ed_le):
            return int(mins), str(tag)
    return 0, "未命中規則"


def _subtract_exclusions(s_dt: pd.Timestamp, e_dt: pd.Timestamp, exclude_ranges):
    if s_dt >= e_dt or not exclude_ranges:
        return [(s_dt, e_dt)]
    segments = [(s_dt, e_dt)]
    for ex_s_t, ex_e_t in exclude_ranges:
        ex_s = pd.Timestamp.combine(s_dt.date(), ex_s_t)
        ex_e = pd.Timestamp.combine(s_dt.date(), ex_e_t)
        new_segments = []
        for a, b in segments:
            if b <= ex_s or a >= ex_e:
                new_segments.append((a, b))
            else:
                if a < ex_s:
                    new_segments.append((a, ex_s))
                if b > ex_e:
                    new_segments.append((ex_e, b))
        segments = [(x, y) for (x, y) in new_segments if x < y]
    return segments


def _coerce_dt_series(series_dt: pd.Series) -> pd.Series:
    if series_dt is None:
        return pd.Series([], dtype="datetime64[ns]")
    return pd.to_datetime(series_dt, errors="coerce").dropna()


def _clamp_first(first_dt: pd.Timestamp, last_dt: pd.Timestamp, clamp_dt: Optional[pd.Timestamp]) -> pd.Timestamp:
    if clamp_dt is None or pd.isna(first_dt) or pd.isna(last_dt):
        return first_dt
    if (first_dt < clamp_dt) and (clamp_dt <= last_dt):
        return clamp_dt
    return first_dt


def _compute_idle(
    series_dt: pd.Series,
    min_minutes: int,
    exclude_ranges: List[Tuple[dt.time, dt.time]],
    clamp_dt: Optional[pd.Timestamp] = None,
) -> Tuple[int, str]:
    s = _coerce_dt_series(series_dt).sort_values()
    if s.size < 2:
        return 0, ""

    if clamp_dt is not None and pd.notna(clamp_dt):
        s2 = s[s >= clamp_dt].copy()
        if s2.empty:
            return 0, ""
        if s2.iloc[0] != clamp_dt:
            s2 = pd.concat([pd.Series([clamp_dt]), s2], ignore_index=True)
        s = s2.sort_values()

    total_min, ranges_txt = 0, []
    prev = s.iloc[0]
    for cur in s.iloc[1:]:
        if cur <= prev:
            prev = cur
            continue
        for a, b in _subtract_exclusions(prev, cur, exclude_ranges or []):
            gap_min = int(round((b - a).total_seconds() / 60.0))
            if gap_min >= int(min_minutes):
                total_min += gap_min
                ranges_txt.append(f"{a.time()} ~ {b.time()}")
        prev = cur

    return int(total_min), "；".join(ranges_txt)


def _span_metrics(series_dt: pd.Series) -> Tuple[pd.Timestamp, pd.Timestamp, int]:
    s = _coerce_dt_series(series_dt)
    if s.empty:
        return pd.NaT, pd.NaT, 0
    return s.min(), s.max(), int(s.size)


def _eff(n: int, m_minutes: int) -> float:
    return round((n / m_minutes * 60.0), 2) if m_minutes and m_minutes > 0 else 0.0


def compute_am_pm_for_group(
    g: pd.DataFrame,
    idle_threshold_min: int,
    exclude_idle_ranges: List[Tuple[dt.time, dt.time]],
    start_time: Optional[dt.time] = None,
) -> pd.Series:
    times = _coerce_dt_series(g["__dt__"])
    if times.empty:
        return pd.Series({
            "第一筆時間": pd.NaT, "最後一筆時間": pd.NaT, "當日筆數": 0,
            "休息分鐘_整體": 0, "命中規則": "無時間資料",
            "當日工時_分鐘_扣休": 0, "效率_件每小時": 0.0,

            "上午_第一筆": pd.NaT, "上午_最後一筆": pd.NaT, "上午_筆數": 0,
            "上午_工時_分鐘": 0, "上午_效率_件每小時": 0.0,
            "上午_空窗分鐘": 0, "上午_空窗時段": "",

            "下午_第一筆": pd.NaT, "下午_最後一筆": pd.NaT, "下午_筆數": 0,
            "下午_休息分鐘": 0, "下午_命中規則": "無時間資料",
            "下午_工時_分鐘_扣休": 0, "下午_效率_件每小時": 0.0,
            "下午_空窗分鐘_扣休": 0, "下午_空窗時段": "",

            "比對棚別筆數": 0, "比對棚別率": 0.0,
        })

    shelf_match = g.get("__shelf_match__", pd.Series([False] * len(g), index=g.index))
    shelf_match = shelf_match.fillna(False).astype(bool)
    match_whole = int(shelf_match.sum())

    clamp_dt_whole: Optional[pd.Timestamp] = None
    if isinstance(start_time, dt.time):
        clamp_dt_whole = pd.Timestamp.combine(times.min().date(), start_time)

    # 上午
    t_am = times[times.dt.time.between(AM_START, AM_END)]
    am_first, am_last, am_cnt = _span_metrics(t_am)
    clamp_dt_am = None
    if clamp_dt_whole is not None and pd.notna(am_first) and pd.notna(am_last):
        if AM_START <= clamp_dt_whole.time() <= AM_END:
            clamp_dt_am = clamp_dt_whole

    am_first_adj = _clamp_first(am_first, am_last, clamp_dt_am)
    am_mins = int(round(((am_last - am_first_adj).total_seconds() / 60.0))) if am_cnt > 0 else 0
    am_mins = max(am_mins, 0)
    am_eff = _eff(am_cnt, am_mins)
    am_idle_min, am_idle_ranges = _compute_idle(
        t_am,
        min_minutes=int(idle_threshold_min),
        exclude_ranges=exclude_idle_ranges,
        clamp_dt=am_first_adj if (am_cnt > 0 and pd.notna(am_first_adj)) else None,
    )

    # 下午
    t_pm = times[times.dt.time.between(PM_START, PM_END)]
    pm_first, pm_last, pm_cnt = _span_metrics(t_pm)
    clamp_dt_pm = None
    if clamp_dt_whole is not None and pd.notna(pm_first) and pd.notna(pm_last):
        if PM_START <= clamp_dt_whole.time() <= PM_END:
            clamp_dt_pm = clamp_dt_whole

    pm_first_adj = _clamp_first(pm_first, pm_last, clamp_dt_pm)

    if pm_cnt > 0 and pd.notna(pm_first_adj) and pd.notna(pm_last):
        pm_break, pm_rule = break_minutes_for_span(pm_first_adj, pm_last)
        raw_pm_mins = (pm_last - pm_first_adj).total_seconds() / 60.0
        pm_mins = max(int(round(raw_pm_mins - pm_break)), 0)
    else:
        pm_break, pm_rule, pm_mins = 0, "無時間資料", 0

    pm_eff = _eff(pm_cnt, pm_mins)
    pm_idle_min, pm_idle_ranges = _compute_idle(
        t_pm,
        min_minutes=int(idle_threshold_min),
        exclude_ranges=exclude_idle_ranges,
        clamp_dt=pm_first_adj if (pm_cnt > 0 and pd.notna(pm_first_adj)) else None,
    )

    # 整體
    whole_first, whole_last, day_cnt = _span_metrics(times)
    whole_first_adj = _clamp_first(whole_first, whole_last, clamp_dt_whole)

    if day_cnt > 0 and pd.notna(whole_first_adj) and pd.notna(whole_last):
        whole_break, br_tag_whole = break_minutes_for_span(whole_first_adj, whole_last)
        raw_whole_mins = (whole_last - whole_first_adj).total_seconds() / 60.0
        whole_mins = max(int(round(raw_whole_mins - whole_break)), 0)
    else:
        whole_break, br_tag_whole, whole_mins = 0, "無時間資料", 0

    whole_eff = _eff(day_cnt, whole_mins)
    match_rate_whole = (match_whole / int(day_cnt)) if int(day_cnt) > 0 else 0.0

    return pd.Series({
        "第一筆時間": whole_first_adj, "最後一筆時間": whole_last, "當日筆數": int(day_cnt),
        "休息分鐘_整體": int(whole_break), "命中規則": br_tag_whole,
        "當日工時_分鐘_扣休": int(whole_mins), "效率_件每小時": whole_eff,

        "上午_第一筆": am_first_adj, "上午_最後一筆": am_last, "上午_筆數": int(am_cnt),
        "上午_工時_分鐘": int(am_mins), "上午_效率_件每小時": am_eff,
        "上午_空窗分鐘": int(am_idle_min), "上午_空窗時段": am_idle_ranges,

        "下午_第一筆": pm_first_adj, "下午_最後一筆": pm_last, "下午_筆數": int(pm_cnt),
        "下午_休息分鐘": int(pm_break), "下午_命中規則": pm_rule,
        "下午_工時_分鐘_扣休": int(pm_mins), "下午_效率_件每小時": pm_eff,
        "下午_空窗分鐘_扣休": int(pm_idle_min), "下午_空窗時段": pm_idle_ranges,

        "比對棚別筆數": int(match_whole),
        "比對棚別率": float(match_rate_whole),
    })


# =========================================================
# Excel 匯出（bytes）
# =========================================================
def autosize_columns(ws, df: pd.DataFrame):
    from openpyxl.utils import get_column_letter
    cols = list(df.columns) if df is not None else []
    for i, col in enumerate(cols, start=1):
        if df is not None and not df.empty:
            sample = [len(str(x)) for x in df[col].head(800).tolist()]
            max_len = max([len(str(col))] + sample)
        else:
            max_len = max(len(str(col)), 8)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)


def shade_rows_by_efficiency(ws, header_name="效率_件每小時", green="C6EFCE", red="FFC7CE", target_eff=20):
    """
    彙總/明細：整列底色
    - val >= target_eff -> green
    - val <  target_eff -> red
    （這段不含字色，保持你原先邏輯；你要的「總表整列底色+字色」已在 _write_total_sheet 內處理）
    """
    from openpyxl.styles import PatternFill

    eff_col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(row=1, column=c).value).strip() == header_name:
            eff_col = c
            break
    if eff_col is None:
        return

    green_fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
    red_fill = PatternFill(start_color=red, end_color=red, fill_type="solid")

    thr = float(target_eff)  # ✅ 修正：floatfloat -> float

    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=eff_col).value
        try:
            val = float(v) if v is not None and str(v).strip() != "" else None
        except Exception:
            val = None
        if val is None:
            continue

        fill = green_fill if val >= thr else red_fill
        for cc in range(1, ws.max_column + 1):
            ws.cell(row=r, column=cc).fill = fill


def _fmt_ts_time(x: Any) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    if isinstance(x, pd.Timestamp):
        if pd.isna(x):
            return ""
        return x.strftime("%H:%M:%S")
    try:
        xx = pd.to_datetime(x, errors="coerce")
        if pd.isna(xx):
            return ""
        return xx.strftime("%H:%M:%S")
    except Exception:
        return ""


def _build_shift_total_df(daily: pd.DataFrame, user_col: str, shift: str) -> pd.DataFrame:
    """
    shift='AM' or 'PM'
    欄位：
    代碼、姓名、筆數、工作區間、總分鐘、效率(件/時)、休息分鐘、空窗分鐘、空窗時段
    """
    if daily is None or daily.empty:
        return pd.DataFrame(columns=["代碼", "姓名", "筆數", "工作區間", "總分鐘", "效率(件/時)", "休息分鐘", "空窗分鐘", "空窗時段"])

    d = daily.copy()

    if shift.upper() == "AM":
        cnt_col = "上午_筆數"
        first_col, last_col = "上午_第一筆", "上午_最後一筆"
        mins_col = "上午_工時_分鐘"
        eff_col = "上午_效率_件每小時"
        rest_col = None
        idle_min_col = "上午_空窗分鐘"
        idle_rng_col = "上午_空窗時段"
    else:
        cnt_col = "下午_筆數"
        first_col, last_col = "下午_第一筆", "下午_最後一筆"
        mins_col = "下午_工時_分鐘_扣休"
        eff_col = "下午_效率_件每小時"
        rest_col = "下午_休息分鐘"
        idle_min_col = "下午_空窗分鐘_扣休"
        idle_rng_col = "下午_空窗時段"

    d[cnt_col] = pd.to_numeric(d.get(cnt_col, 0), errors="coerce").fillna(0).astype(int)
    d = d[d[cnt_col] > 0].copy()

    if d.empty:
        return pd.DataFrame(columns=["代碼", "姓名", "筆數", "工作區間", "總分鐘", "效率(件/時)", "休息分鐘", "空窗分鐘", "空窗時段"])

    name_series = d["對應姓名"].astype(str).fillna("").str.strip()
    code_series = d[user_col].astype(str).fillna("").str.strip()
    d["_姓名顯示"] = name_series.where(name_series.ne(""), code_series)

    d["工作區間"] = d.apply(lambda r: f"{_fmt_ts_time(r.get(first_col))} ~ {_fmt_ts_time(r.get(last_col))}".strip(), axis=1)

    out = pd.DataFrame({
        "代碼": code_series,
        "姓名": d["_姓名顯示"],
        "筆數": d[cnt_col].astype(int),
        "工作區間": d["工作區間"],
        "總分鐘": pd.to_numeric(d.get(mins_col, 0), errors="coerce").fillna(0).astype(int),
        "效率(件/時)": pd.to_numeric(d.get(eff_col, 0), errors="coerce").fillna(0.0).round(2),
        "休息分鐘": (pd.to_numeric(d.get(rest_col, 0), errors="coerce").fillna(0).astype(int) if rest_col else 0),
        "空窗分鐘": pd.to_numeric(d.get(idle_min_col, 0), errors="coerce").fillna(0).astype(int),
        "空窗時段": d.get(idle_rng_col, "").astype(str).fillna(""),
    })

    out = out.sort_values(["效率(件/時)", "代碼"], ascending=[False, True]).reset_index(drop=True)
    return out


def _write_total_sheet(ws, daily: pd.DataFrame, user_col: str):
    """
    在同一張「總表」工作表，依日期輸出：
    [YYYY-MM-DD 上架績效] / 上午表 / 下午表

    ✅ 你指定的新規則（整列套色）：
      效率 < 20  → 整列底色 #FFC7CE、整列文字 #9C0006
      效率 >= 20 → 整列底色 #C6EFCE、整列文字 #006100
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    EFF_THRESHOLD = 20.0

    thin = Side(style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_title = PatternFill("solid", fgColor="FFFFFF")
    fill_header = PatternFill("solid", fgColor="E5E7EB")   # 灰
    fill_am = PatternFill("solid", fgColor="D1FAE5")       # 上午淡綠
    fill_pm = PatternFill("solid", fgColor="FDE2E2")       # 下午淡粉

    # ✅ 低效：紅底紅字（效率 < 20）
    eff_low_fill = PatternFill("solid", fgColor="FFC7CE")  # #FFC7CE
    eff_low_font = Font(color="9C0006")                    # #9C0006

    # ✅ 高效：綠底綠字（效率 >= 20）
    eff_high_fill = PatternFill("solid", fgColor="C6EFCE") # #C6EFCE
    eff_high_font = Font(color="006100")                   # #006100

    font_title = Font(bold=True, size=14)
    font_section = Font(bold=True, size=12)
    font_header = Font(bold=True, size=11)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["代碼", "姓名", "筆數", "工作區間", "總分鐘", "效率(件/時)", "休息分鐘", "空窗分鐘", "空窗時段"]
    ncol = len(headers)

    # 欄寬
    col_widths = [12, 10, 6, 22, 8, 10, 8, 8, 60]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if daily is None or daily.empty or "日期" not in daily.columns:
        ws["A1"] = "無可用資料"
        return

    dates = sorted([x for x in daily["日期"].dropna().unique()])
    r = 1

    def _try_float(x):
        try:
            if x is None:
                return None
            s = str(x).strip()
            if s == "":
                return None
            return float(s)
        except Exception:
            return None

    for d0 in dates:
        day_df = daily[daily["日期"] == d0].copy()

        # Title
        title = f"{d0} 上架績效"
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        c = ws.cell(row=r, column=1, value=title)
        c.fill = fill_title
        c.font = font_title
        c.alignment = align_center
        r += 1

        # ======================
        # AM
        # ======================
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        c = ws.cell(row=r, column=1, value="上午")
        c.font = font_section
        c.alignment = align_center
        r += 1

        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=j, value=h)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border
        r += 1

        am_tbl = _build_shift_total_df(day_df, user_col=user_col, shift="AM")
        if am_tbl.empty:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            c = ws.cell(row=r, column=1, value="（上午無資料）")
            c.alignment = align_center
            r += 1
        else:
            for _, row in am_tbl.iterrows():
                eff_val = _try_float(row.get("效率(件/時)", ""))

                # 預設：上午底色；若有效率值→整列依效率規則套色
                row_fill = fill_am
                row_font = None
                if eff_val is not None:
                    if eff_val < EFF_THRESHOLD:
                        row_fill = eff_low_fill
                        row_font = eff_low_font
                    else:
                        row_fill = eff_high_fill
                        row_font = eff_high_font

                for j, h in enumerate(headers, start=1):
                    v = row.get(h, "")
                    cell = ws.cell(row=r, column=j, value=v)
                    cell.fill = row_fill
                    if row_font is not None:
                        cell.font = row_font
                    cell.alignment = (align_left if h == "空窗時段" else align_center)
                    cell.border = border

                r += 1

        r += 1  # blank

        # ======================
        # PM
        # ======================
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        c = ws.cell(row=r, column=1, value="下午")
        c.font = font_section
        c.alignment = align_center
        r += 1

        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=j, value=h)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border
        r += 1

        pm_tbl = _build_shift_total_df(day_df, user_col=user_col, shift="PM")
        if pm_tbl.empty:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            c = ws.cell(row=r, column=1, value="（下午無資料）")
            c.alignment = align_center
            r += 1
        else:
            for _, row in pm_tbl.iterrows():
                eff_val = _try_float(row.get("效率(件/時)", ""))

                row_fill = fill_pm
                row_font = None
                if eff_val is not None:
                    if eff_val < EFF_THRESHOLD:
                        row_fill = eff_low_fill
                        row_font = eff_low_font
                    else:
                        row_fill = eff_high_fill
                        row_font = eff_high_font

                for j, h in enumerate(headers, start=1):
                    v = row.get(h, "")
                    cell = ws.cell(row=r, column=j, value=v)
                    cell.fill = row_fill
                    if row_font is not None:
                        cell.font = row_font
                    cell.alignment = (align_left if h == "空窗時段" else align_center)
                    cell.border = border

                r += 1

        r += 2  # blank between dates

    ws.freeze_panes = "A4"


def build_excel_bytes(
    user_col: str,
    summary_out: pd.DataFrame,
    daily: pd.DataFrame,
    shelf_person_pivot: pd.DataFrame,
    stype_person_pivot: pd.DataFrame,
    target_eff: float,
) -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm:ss", date_format="yyyy-mm-dd") as writer:
        sum_cols = [
            user_col, "對應姓名", "総日數",
            "總筆數", "總工時_分鐘_扣休", "效率_件每小時",
            "上午筆數", "上午工時_分鐘", "上午效率_件每小時",
            "下午筆數", "下午工時_分鐘_扣休", "下午效率_件每小時",
            "比對棚別筆數", "比對棚別率",
        ]
        summary_out[sum_cols].to_excel(writer, index=False, sheet_name="彙總")
        autosize_columns(writer.sheets["彙總"], summary_out[sum_cols])
        # 這裡沿用舊邏輯：>=target 綠、<target 紅（不影響你要的總表）
        shade_rows_by_efficiency(writer.sheets["彙總"], "效率_件每小時", target_eff=target_eff)

        det_cols = [
            user_col, "對應姓名", "日期",
            "第一筆時間", "最後一筆時間", "當日筆數",
            "休息分鐘_整體", "當日工時_分鐘_扣休", "效率_件每小時",
            "上午_第一筆", "上午_最後一筆", "上午_筆數", "上午_工時_分鐘", "上午_效率_件每小時",
            "上午_空窗分鐘", "上午_空窗時段",
            "下午_第一筆", "下午_最後一筆", "下午_筆數", "下午_休息分鐘",
            "下午_工時_分鐘_扣休", "下午_效率_件每小時",
            "下午_空窗分鐘_扣休", "下午_空窗時段",
            "比對棚別筆數", "比對棚別率",
        ]
        daily.sort_values([user_col, "日期", "第一筆時間"])[det_cols].to_excel(writer, index=False, sheet_name="明細")
        autosize_columns(writer.sheets["明細"], daily[det_cols])
        shade_rows_by_efficiency(writer.sheets["明細"], "效率_件每小時", target_eff=target_eff)

        if stype_person_pivot is not None and not stype_person_pivot.empty:
            stype_person_pivot.to_excel(writer, index=False, sheet_name="儲位類型_人員_樞紐")
            autosize_columns(writer.sheets["儲位類型_人員_樞紐"], stype_person_pivot)

        if shelf_person_pivot is not None and not shelf_person_pivot.empty:
            shelf_person_pivot.to_excel(writer, index=False, sheet_name="棚別_人員_樞紐")
            autosize_columns(writer.sheets["棚別_人員_樞紐"], shelf_person_pivot)

        rules_rows = []
        for i, (st_ge, ed_le, mins, tag) in enumerate(BREAK_RULES, start=1):
            rules_rows.append({
                "優先序": i,
                "首時間條件(>=)": st_ge.strftime("%H:%M:%S"),
                "末時間條件(<=)": ed_le.strftime("%H:%M:%S"),
                "休息分鐘": int(mins),
                "規則說明": str(tag),
            })
        rules_df = pd.DataFrame(rules_rows)
        rules_df.to_excel(writer, index=False, sheet_name="休息規則")
        autosize_columns(writer.sheets["休息規則"], rules_df)

        # ✅ 新增：總表（你要的整列套色在這張）
        ws_total = writer.book.create_sheet("總表")
        writer.sheets["總表"] = ws_total
        _write_total_sheet(ws_total, daily=daily, user_col=user_col)

    return out.getvalue()


# =========================================================
# Streamlit Page
# =========================================================
def main():
    inject_logistics_theme()
    set_page(
        "上架產能分析（Putaway KPI）",
        icon="📦",
        subtitle="主畫面顯示：儲位類型樞紐表 + 棚別樞紐表；Excel 匯出含：彙總/明細/總表"
    )

    if "putaway_last" not in st.session_state:
        st.session_state.putaway_last = None

    # Sidebar：上架人設定
    render_putaway_people_settings_panel()

    # Sidebar：統一條件
    controls = sidebar_controls(default_top_n=30, enable_exclude_windows=True, state_key_prefix="putaway")
    top_n = int(controls.get("top_n", 30))

    exclude_raw = _extract_exclude_value_from_controls(controls)
    exclude_idle_ranges = _parse_exclude_windows(exclude_raw)

    with st.sidebar:
        st.markdown("---")
        target_eff = st.number_input("達標門檻（效率 ≥）", min_value=1, max_value=999, value=int(TARGET_EFF_DEFAULT), step=1)
        idle_threshold = st.number_input("空窗門檻（分鐘 ≥ 才算）", min_value=1, max_value=240, value=int(IDLE_MIN_THRESHOLD_DEFAULT), step=1)

        start_time_raw = st.text_input("起始時間（clamp 第一筆；HH:MM:SS）", value="08:05:00", key="putaway_global_start_time")
        global_start_time = _parse_time_any(start_time_raw)
        if global_start_time is None:
            st.warning("⚠️ 起始時間格式不正確，將不進行 clamp（請用 HH:MM 或 HH:MM:SS）")
        else:
            st.caption(f"✅ clamp 起始時間：{global_start_time.strftime('%H:%M:%S')}")

        preview = "、".join([f"{a.strftime('%H:%M')}~{b.strftime('%H:%M')}" for a, b in exclude_idle_ranges]) if exclude_idle_ranges else "（無）"
        st.caption(f"✅ 已讀取排除空窗時段：{preview}")
        st.caption("⚠️ 若你改了條件/棚別主檔，需再按一次「🚀 產出 KPI」才會重新計算。")
        st.caption("提示：上傳 .xls 需 requirements 安裝 xlrd==2.0.1")

    # ✅ 棚別主檔在主畫面（只上傳，不展示表格）
    card_open("📚 棚別主檔（儲位明細）")
    slot_master_file = st.file_uploader(
        "上傳棚別主檔（需含欄位：『儲位』『棚別』）",
        type=["xlsx", "xlsm", "xls", "csv"],
        key="putaway_slot_master_main",
        help="比對：明細欄位『到』(儲位) → 主檔『儲位』對應出『棚別』；查得到棚別即算比對成功。",
    )
    card_close()

    # 上傳作業原始資料
    card_open("📤 上傳作業原始資料（上架）")
    uploaded = st.file_uploader(
        "上傳 Excel / CSV（需包含：由、到、修訂日期/時間、記錄輸入人）",
        type=["xlsx", "xlsm", "xls", "csv"],
        label_visibility="collapsed",
        key="putaway_raw",
    )
    run_clicked = st.button("🚀 產出 KPI", type="primary", disabled=uploaded is None)
    card_close()

    # ✅ 條件變更提醒
    last = st.session_state.putaway_last
    people_settings_snapshot = _get_putaway_people_settings()
    people_hash = str(sorted([(k, v.get("name", ""), v.get("area", "")) for k, v in people_settings_snapshot.items()]))

    slot_hash = ""
    if slot_master_file is not None:
        slot_hash = f"{slot_master_file.name}:{len(slot_master_file.getvalue())}"

    current_params = {
        "target_eff": int(target_eff),
        "idle_threshold": int(idle_threshold),
        "exclude_idle_ranges": [(a.strftime("%H:%M:%S"), b.strftime("%H:%M:%S")) for a, b in exclude_idle_ranges],
        "top_n": int(top_n),
        "people_hash": people_hash,
        "global_start_time": global_start_time.strftime("%H:%M:%S") if global_start_time else "",
        "slot_hash": slot_hash,
    }
    if last and last.get("params") and last.get("params") != current_params:
        st.warning("⚠️ 你已變更條件（含棚別主檔），請再按一次「🚀 產出 KPI」才會套用新條件。")

    # ✅ 計算
    if run_clicked:
        with st.spinner("計算中，請稍候..."):
            sheets = read_excel_any_quiet_bytes(uploaded.name, uploaded.getvalue())

            kept_all = []
            for sn, df in sheets.items():
                k = prepare_filtered_df(df)
                if not k.empty:
                    k["__sheet__"] = sn
                    kept_all.append(k)

            if not kept_all:
                st.error("無符合資料（可能缺『由/到』欄或過濾後為空）。")
                st.session_state.putaway_last = None
                return

            data = pd.concat(kept_all, ignore_index=True)

            user_col = find_first_column(data, INPUT_USER_CANDIDATES)
            revdt_col = find_first_column(data, REV_DT_CANDIDATES)
            if user_col is None:
                st.error("找不到『記錄輸入人』欄位（候選：記錄輸入人/記錄輸入者/建立人/輸入人）。")
                st.session_state.putaway_last = None
                return
            if revdt_col is None:
                st.error("找不到『修訂日期/時間』欄位（候選：修訂日期/修訂時間/修訂日/異動時間/修改時間）。")
                st.session_state.putaway_last = None
                return

            data["__dt__"] = pd.to_datetime(data[revdt_col], errors="coerce")
            data["__code__"] = data[user_col].astype(str).str.strip()

            # ✅ 上架人姓名
            people_settings = _get_putaway_people_settings()
            custom_name_map = {k: v.get("name", "") for k, v in people_settings.items() if v.get("name")}
            merged_name_map = {**NAME_MAP, **custom_name_map}
            data["對應姓名"] = data["__code__"].map(merged_name_map).fillna("")

            # ✅ 棚別比對（到→棚別）
            data["__to_loc__"] = data["到"].astype(str).str.strip()

            slot_map_shelf = {}
            if slot_master_file is not None:
                try:
                    slot_master_df = load_slot_master_bytes(slot_master_file.name, slot_master_file.getvalue())
                    if not slot_master_df.empty:
                        slot_map_shelf = dict(zip(slot_master_df["儲位"], slot_master_df["棚別"]))
                    else:
                        st.warning("⚠️ 棚別主檔讀取成功但缺必要欄位（需含：儲位、棚別），將不計算棚別。")
                except Exception as e:
                    st.warning(f"⚠️ 棚別主檔讀取失敗：{e}，將不計算棚別。")

            if slot_map_shelf:
                data["棚別"] = data["__to_loc__"].map(slot_map_shelf).fillna("")
            else:
                data["棚別"] = ""

            data["__shelf_match__"] = data["棚別"].astype(str).str.strip().ne("")

            # ✅ 儲位類型：棚別抓區碼3，抓不到用 到(儲位)
            data["棚別_區碼3"] = data["棚別"].apply(_extract_zone3)
            fallback_zone3 = data["__to_loc__"].apply(_extract_zone3)
            data["棚別_區碼3"] = data["棚別_區碼3"].where(data["棚別_區碼3"].ne(""), fallback_zone3)
            data["儲位類型"] = data["棚別_區碼3"].apply(_map_storage_type).fillna("")

            dt_data = data.dropna(subset=["__dt__"]).copy()
            if dt_data.empty:
                st.error("資料沒有可用的修訂日期時間，無法計算。")
                st.session_state.putaway_last = None
                return

            dt_data["日期"] = dt_data["__dt__"].dt.date

            # ✅ 日彙總
            daily = (
                dt_data.groupby([user_col, "對應姓名", "日期"], dropna=False)
                .apply(lambda g: compute_am_pm_for_group(
                    g,
                    idle_threshold_min=int(idle_threshold),
                    exclude_idle_ranges=exclude_idle_ranges,
                    start_time=global_start_time,
                ))
                .reset_index()
            )

            # ✅ 個人總彙總
            summary = (
                daily.groupby([user_col, "對應姓名"], dropna=False, as_index=False)
                .agg(
                    総日數=("日期", "nunique"),
                    總筆數=("當日筆數", "sum"),
                    總工時_分鐘_扣休=("當日工時_分鐘_扣休", "sum"),
                    上午筆數=("上午_筆數", "sum"),
                    上午工時_分鐘=("上午_工時_分鐘", "sum"),
                    下午筆數=("下午_筆數", "sum"),
                    下午工時_分鐘_扣休=("下午_工時_分鐘_扣休", "sum"),
                    比對棚別筆數=("比對棚別筆數", "sum"),
                )
            )

            summary["上午效率_件每小時"] = summary.apply(lambda r: _eff(int(r["上午筆數"]), int(r["上午工時_分鐘"])), axis=1)
            summary["下午效率_件每小時"] = summary.apply(lambda r: _eff(int(r["下午筆數"]), int(r["下午工時_分鐘_扣休"])), axis=1)
            summary["總工時_分鐘_扣休"] = summary["上午工時_分鐘"].fillna(0).astype(int) + summary["下午工時_分鐘_扣休"].fillna(0).astype(int)
            summary["效率_件每小時"] = summary.apply(lambda r: _eff(int(r["總筆數"]), int(r["總工時_分鐘_扣休"])), axis=1)

            for c in ["總筆數", "總工時_分鐘_扣休", "上午筆數", "上午工時_分鐘", "下午筆數", "下午工時_分鐘_扣休", "比對棚別筆數"]:
                summary[c] = summary[c].fillna(0).astype(int)

            summary["比對棚別率"] = summary.apply(
                lambda r: (int(r["比對棚別筆數"]) / int(r["總筆數"])) if int(r["總筆數"]) > 0 else 0.0,
                axis=1,
            )

            total_people = int(summary[user_col].nunique())
            met_people = int((summary["效率_件每小時"] >= float(target_eff)).sum())
            rate = (met_people / total_people) if total_people > 0 else 0.0

            total_match = int(summary["比對棚別筆數"].sum())
            total_cnt = int(summary["總筆數"].sum())
            match_rate_all = (total_match / total_cnt) if total_cnt > 0 else 0.0

            total_row = {
                user_col: "整體合計",
                "對應姓名": "",
                "総日數": int(summary["総日數"].sum()),
                "總筆數": int(summary["總筆數"].sum()),
                "總工時_分鐘_扣休": int(summary["總工時_分鐘_扣休"].sum()),
                "上午筆數": int(summary["上午筆數"].sum()),
                "上午工時_分鐘": int(summary["上午工時_分鐘"].sum()),
                "下午筆數": int(summary["下午筆數"].sum()),
                "下午工時_分鐘_扣休": int(summary["下午工時_分鐘_扣休"].sum()),
                "效率_件每小時": _eff(int(summary["總筆數"].sum()), int(summary["總工時_分鐘_扣休"].sum())),
                "上午效率_件每小時": _eff(int(summary["上午筆數"].sum()), int(summary["上午工時_分鐘"].sum())),
                "下午效率_件每小時": _eff(int(summary["下午筆數"].sum()), int(summary["下午工時_分鐘_扣休"].sum())),
                "比對棚別筆數": int(total_match),
                "比對棚別率": float(match_rate_all),
            }
            summary_out = pd.concat([summary, pd.DataFrame([total_row])], ignore_index=True)

            # ✅ 樞紐：每人每棚別
            shelf_for_group = dt_data["棚別"].astype(str).str.strip()
            shelf_for_group = shelf_for_group.where(shelf_for_group.ne(""), "未比對")
            shelf_person_long = (
                dt_data.assign(_棚別分類=shelf_for_group)
                .groupby([user_col, "對應姓名", "_棚別分類"], dropna=False)
                .size()
                .reset_index(name="筆數")
                .rename(columns={"_棚別分類": "棚別"})
            )
            if not shelf_person_long.empty:
                shelf_person_pivot = (
                    shelf_person_long.pivot_table(
                        index=[user_col, "對應姓名"],
                        columns="棚別",
                        values="筆數",
                        aggfunc="sum",
                        fill_value=0,
                    )
                    .reset_index()
                )
                cols = [c for c in shelf_person_pivot.columns if c not in (user_col, "對應姓名")]
                cols_sorted = [c for c in cols if c != "未比對"] + (["未比對"] if "未比對" in cols else [])
                shelf_person_pivot = shelf_person_pivot[[user_col, "對應姓名"] + cols_sorted]
            else:
                shelf_person_pivot = pd.DataFrame()

            # ✅ 樞紐：每人每儲位類型
            stype_for_group = dt_data["儲位類型"].astype(str).str.strip()
            stype_for_group = stype_for_group.where(stype_for_group.ne(""), "未分類")
            stype_person_long = (
                dt_data.assign(_儲位類型分類=stype_for_group)
                .groupby([user_col, "對應姓名", "_儲位類型分類"], dropna=False)
                .size()
                .reset_index(name="筆數")
                .rename(columns={"_儲位類型分類": "儲位類型"})
            )
            if not stype_person_long.empty:
                stype_person_pivot = (
                    stype_person_long.pivot_table(
                        index=[user_col, "對應姓名"],
                        columns="儲位類型",
                        values="筆數",
                        aggfunc="sum",
                        fill_value=0,
                    )
                    .reset_index()
                )
                prefer = ["輕型料架", "重型低空", "落地儲", "高空儲", "未分類"]
                cols = [c for c in stype_person_pivot.columns if c not in (user_col, "對應姓名")]
                ordered = [c for c in prefer if c in cols] + [c for c in cols if c not in prefer]
                stype_person_pivot = stype_person_pivot[[user_col, "對應姓名"] + ordered]
            else:
                stype_person_pivot = pd.DataFrame()

            xlsx_bytes = build_excel_bytes(
                user_col=user_col,
                summary_out=summary_out,
                daily=daily,
                shelf_person_pivot=shelf_person_pivot,
                stype_person_pivot=stype_person_pivot,
                target_eff=float(target_eff),
            )
            xlsx_name = f"{uploaded.name.rsplit('.', 1)[0]}_上架績效.xlsx"

            st.session_state.putaway_last = {
                "params": current_params,
                "user_col": user_col,
                "summary": summary,
                "target_eff": float(target_eff),
                "top_n": int(top_n),
                "total_people": int(total_people),
                "met_people": int(met_people),
                "rate": float(rate),
                "xlsx_bytes": xlsx_bytes,
                "xlsx_name": xlsx_name,
                "total_match": int(total_match),
                "match_rate_all": float(match_rate_all),
                "shelf_person_pivot": shelf_person_pivot,
                "stype_person_pivot": stype_person_pivot,
            }

    # ======================
    # ✅ 顯示（主畫面：只顯示兩個樞紐表）
    # ======================
    last = st.session_state.putaway_last
    if not last:
        st.info("請先上傳上架作業原始資料並點選「🚀 產出 KPI」")
        return

    user_col = last["user_col"]
    summary = last["summary"]
    target_eff_show = float(last["target_eff"])
    top_n_show = int(controls.get("top_n", last.get("top_n", 30)))
    total_people = int(last["total_people"])
    met_people = int(last["met_people"])
    rate = float(last["rate"])
    xlsx_bytes = last["xlsx_bytes"]
    xlsx_name = last["xlsx_name"]
    total_match = int(last.get("total_match", 0))
    match_rate_all = float(last.get("match_rate_all", 0.0))

    shelf_person_pivot = last.get("shelf_person_pivot", pd.DataFrame())
    stype_person_pivot = last.get("stype_person_pivot", pd.DataFrame())

    # KPI（不是表格）
    card_open("📌 總覽 KPI")
    render_kpis([
        KPI("總人數", f"{total_people:,}"),
        KPI("達標人數", f"{met_people:,}"),
        KPI("達標率", f"{rate:.1%}"),
        KPI("達標門檻", f"效率 ≥ {int(target_eff_show)}"),
        KPI("棚別比對筆數", f"{total_match:,}"),
        KPI("棚別比對率", f"{match_rate_all:.1%}"),
    ])
    card_close()

    # ✅ 只顯示兩張表：儲位類型樞紐 + 棚別樞紐
    card_open("📦 樞紐表（每人一列、每儲位類型一欄）")
    if stype_person_pivot is None or stype_person_pivot.empty:
        st.info("尚未產生儲位類型樞紐表（可能無法擷取區碼3或資料為空）。")
    else:
        st.dataframe(stype_person_pivot, use_container_width=True, hide_index=True)
    card_close()

    card_open("🏷️ 樞紐表（每人一列、每棚別一欄）")
    if shelf_person_pivot is None or shelf_person_pivot.empty:
        st.info("尚未產生棚別主檔樞紐表（可能未上傳棚別主檔，或比對結果為空）。")
    else:
        st.dataframe(shelf_person_pivot, use_container_width=True, hide_index=True)
    card_close()

    # AM/PM 圖表（不是表格）
    col_l, col_r = st.columns(2)

    with col_l:
        card_open(f"🌓 AM（上午）效率排行（Top {top_n_show}）")
        am_rank = summary[[user_col, "對應姓名", "上午筆數", "上午工時_分鐘", "上午效率_件每小時"]].copy()
        am_rank = am_rank.rename(columns={"上午效率_件每小時": "效率", "上午筆數": "筆數", "上午工時_分鐘": "工時"})
        am_rank["姓名"] = am_rank["對應姓名"].where(am_rank["對應姓名"].astype(str).str.len() > 0, am_rank[user_col].astype(str))
        bar_topN(
            am_rank[["姓名", "效率", "筆數", "工時"]],
            x_col="姓名",
            y_col="效率",
            hover_cols=["筆數", "工時"],
            top_n=top_n_show,
            target=float(target_eff_show),
        )
        card_close()

    with col_r:
        card_open(f"🌙 PM（下午）效率排行（Top {top_n_show}）")
        pm_rank = summary[[user_col, "對應姓名", "下午筆數", "下午工時_分鐘_扣休", "下午效率_件每小時"]].copy()
        pm_rank = pm_rank.rename(columns={"下午效率_件每小時": "效率", "下午筆數": "筆數", "下午工時_分鐘_扣休": "工時"})
        pm_rank["姓名"] = pm_rank["對應姓名"].where(pm_rank["對應姓名"].astype(str).str.len() > 0, pm_rank[user_col].astype(str))
        bar_topN(
            pm_rank[["姓名", "效率", "筆數", "工時"]],
            x_col="姓名",
            y_col="效率",
            hover_cols=["筆數", "工時"],
            top_n=top_n_show,
            target=float(target_eff_show),
        )
        card_close()

    download_excel_card(
        xlsx_bytes,
        xlsx_name,
        label="⬇️ 匯出 KPI 報表（Excel：含『總表』）",
    )


if __name__ == "__main__":
    main()
